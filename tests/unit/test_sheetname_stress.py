from __future__ import annotations

import importlib
from pathlib import Path

import pandas as pd
import pytest

from app.infra.io_utils import write_xlsx_atomic  # noqa: E402

_HAS_OPENPYXL = importlib.util.find_spec("openpyxl") is not None


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="openpyxl لازم است برای خواندن .xlsx")
def test_write_xlsx_atomic_handles_many_long_duplicate_names(tmp_path: Path) -> None:
    base = "X" * 40 + ":" + "Y" * 40
    sheets = {f"{base}_{i}": pd.DataFrame({"val": [i]}) for i in range(1, 15)}
    out = tmp_path / "stress.xlsx"

    write_xlsx_atomic(sheets, out)

    from openpyxl import load_workbook  # noqa: WPS433 (local optional import)

    names = load_workbook(out).sheetnames

    assert len(names) == 14
    assert len(names) == len(set(names))
    assert all(len(name) <= 31 for name in names)
