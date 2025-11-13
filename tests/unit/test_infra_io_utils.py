from __future__ import annotations

import hashlib
import importlib
import sys
import warnings
from pathlib import Path

import pandas as pd
import pytest

# مسیر پروژه برای ایمپورت ماژول
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

import app.infra.io_utils as io_utils  # noqa: E402
from app.infra.io_utils import ALT_CODE_COLUMN, write_xlsx_atomic  # noqa: E402

_HAS_OPENPYXL = importlib.util.find_spec("openpyxl") is not None
_HAS_XLSXWRITER = importlib.util.find_spec("xlsxwriter") is not None
_HAS_ENGINE = _HAS_OPENPYXL or _HAS_XLSXWRITER


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="openpyxl لازم است برای خواندن .xlsx")
def test_write_xlsx_atomic_sanitizes_and_deduplicates(tmp_path: Path) -> None:
    df = pd.DataFrame({"a": [1], "b": [2]})
    out = tmp_path / "out.xlsx"

    write_xlsx_atomic({"Sheet/1": df, "Sheet:1": df, " ": df}, out)

    from openpyxl import load_workbook  # local import (optional dep)

    wb = load_workbook(out)
    names = wb.sheetnames

    assert names[0] == "Sheet 1"
    assert names[1] == "Sheet 1 (2)"
    assert names[2] == "Sheet"


@pytest.mark.skipif(not _HAS_ENGINE, reason="هیچ engine اکسل یافت نشد")
def test_write_xlsx_atomic_replaces_existing_file_atomically(tmp_path: Path) -> None:
    df1 = pd.DataFrame({"val": [1, 2, 3]})
    df2 = pd.DataFrame({"val": [4, 5, 6, 7]})
    out = tmp_path / "nested" / "out.xlsx"

    write_xlsx_atomic({"S1": df1}, out)
    h1 = hashlib.sha256(out.read_bytes()).hexdigest()
    mtime1 = out.stat().st_mtime

    write_xlsx_atomic({"S2": df2}, out)
    h2 = hashlib.sha256(out.read_bytes()).hexdigest()
    mtime2 = out.stat().st_mtime

    assert h1 != h2, "فایل باید به‌صورت کامل جایگزین شود (hash متفاوت)"
    assert mtime2 >= mtime1, "mtime باید جلوتر یا برابر باشد (جایگزینی)"


@pytest.mark.skipif(not (_HAS_ENGINE and _HAS_OPENPYXL), reason="engine/openpyxl لازم است")
def test_write_xlsx_atomic_emits_no_pandas_futurewarnings(tmp_path: Path) -> None:
    df = pd.DataFrame({"x": [1, 2]})
    out = tmp_path / "fw.xlsx"

    with warnings.catch_warnings(record=True) as caught:
        warnings.simplefilter("always", category=FutureWarning)
        write_xlsx_atomic({"Sheet": df}, out)

        from openpyxl import load_workbook  # local import (optional dep)

        wb = load_workbook(out)
        assert "Sheet" in wb.sheetnames

        fw = [w for w in caught if issubclass(w.category, FutureWarning)]
        assert not fw, f"FutureWarnings: {[str(w.message) for w in fw]}"


@pytest.mark.skipif(not _HAS_ENGINE, reason="هیچ engine اکسل یافت نشد")
def test_write_xlsx_atomic_accepts_str_path_and_creates_parent(tmp_path: Path) -> None:
    df = pd.DataFrame({"val": [42]})
    nested = tmp_path / "nested" / "more"
    out = nested / "out.xlsx"

    write_xlsx_atomic({"Sheet": df}, str(out))

    assert out.exists()
    assert out.stat().st_size > 0


@pytest.mark.skipif(not _HAS_ENGINE, reason="هیچ engine اکسل یافت نشد")
def test_write_xlsx_atomic_respects_env_override(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    engine = "openpyxl" if _HAS_OPENPYXL else "xlsxwriter"
    monkeypatch.setenv("EXCEL_ENGINE", engine)
    out = tmp_path / "env.xlsx"

    write_xlsx_atomic({"Sheet": pd.DataFrame({"v": [1]})}, out)

    assert out.exists() and out.stat().st_size > 0


@pytest.mark.skipif(
    not (_HAS_OPENPYXL and _HAS_XLSXWRITER),
    reason="xlsxwriter و openpyxl باید نصب باشند",
)
def test_write_xlsx_atomic_aligns_with_xlsxwriter(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("EXCEL_ENGINE", "xlsxwriter")
    df = pd.DataFrame({"کدرشته": list(range(1, 5))})
    out = tmp_path / "xlsxwriter_font.xlsx"

    write_xlsx_atomic({"Sheet": df}, out, font_name="Tahoma")

    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb["Sheet"]

    for idx in range(1, 5):
        cell = ws.cell(row=idx, column=1)
        assert cell.font.name == "Tahoma"
        assert cell.font.size == 8
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "center"

    tables = list(ws.tables.values())
    assert len(tables) == 1
    table = tables[0]
    assert table.ref == "A1:A5"


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="openpyxl لازم است برای بررسی RTL/فونت")
def test_write_xlsx_atomic_applies_rtl_and_font(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("EXCEL_ENGINE", "openpyxl")
    df = pd.DataFrame({"کدرشته": [101]})
    out = tmp_path / "rtl.xlsx"

    write_xlsx_atomic({"Sheet": df}, out, rtl=True, font_name="Tahoma")

    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb["Sheet"]

    assert ws.sheet_view.rightToLeft is True
    assert ws.cell(row=1, column=1).font.name == "Tahoma"
    assert ws.cell(row=1, column=1).font.size == 8


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="openpyxl لازم است برای بررسی فونت")
def test_write_xlsx_atomic_applies_font_to_all_rows(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setenv("EXCEL_ENGINE", "openpyxl")
    df = pd.DataFrame({"کدرشته": list(range(1, 105))})
    out = tmp_path / "full_font.xlsx"

    write_xlsx_atomic({"Sheet": df}, out, font_name="Tahoma")

    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb["Sheet"]

    target_cell = ws.cell(row=60, column=1)
    assert target_cell.font.name == "Tahoma"
    assert target_cell.font.size == 8
    assert target_cell.alignment.horizontal is None
    assert target_cell.alignment.vertical is None

    tables = list(ws.tables.values())
    assert len(tables) == 1
    table = tables[0]
    assert table.ref == "A1:A105"


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="openpyxl لازم است برای خواندن .xlsx")
def test_read_excel_first_sheet_preserves_alt_code_as_text(tmp_path: Path) -> None:
    data = pd.DataFrame({ALT_CODE_COLUMN: [123456], "value": [1]})
    sample = tmp_path / "sample.xlsx"
    data.to_excel(sample, index=False)

    loaded = io_utils.read_excel_first_sheet(sample)

    assert ALT_CODE_COLUMN in loaded.columns
    assert loaded[ALT_CODE_COLUMN].dtype == object
    assert loaded.loc[0, ALT_CODE_COLUMN] == "123456"


@pytest.mark.skipif(not _HAS_OPENPYXL, reason="openpyxl لازم است برای خواندن .xlsx")
def test_read_crosswalk_workbook_coerces_alt_code_in_all_sheets(tmp_path: Path) -> None:
    groups = pd.DataFrame({ALT_CODE_COLUMN: [111222], "گروه": ["الف"]})
    synonyms = pd.DataFrame({ALT_CODE_COLUMN: [333444], "کد اصلی": ["ب"]})
    sample = tmp_path / "crosswalk.xlsx"

    with pd.ExcelWriter(sample) as writer:
        groups.to_excel(writer, sheet_name="پایه تحصیلی (گروه آزمایشی)", index=False)
        synonyms.to_excel(writer, sheet_name="Synonyms", index=False)

    groups_df, synonyms_df = io_utils.read_crosswalk_workbook(sample)

    assert groups_df[ALT_CODE_COLUMN].dtype == object
    assert groups_df.loc[0, ALT_CODE_COLUMN] == "111222"
    assert synonyms_df is not None
    assert synonyms_df[ALT_CODE_COLUMN].dtype == object
    assert synonyms_df.loc[0, ALT_CODE_COLUMN] == "333444"


def test_write_xlsx_atomic_cleans_up_temp_file_on_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    monkeypatch.setattr(io_utils, "_pick_engine", lambda: "openpyxl")

    class ExplodingWriter:
        def __enter__(self) -> "ExplodingWriter":
            raise RuntimeError("boom")

        def __exit__(self, exc_type, exc, tb) -> bool:  # pragma: no cover - interface
            return False

    monkeypatch.setattr(pd, "ExcelWriter", lambda *args, **kwargs: ExplodingWriter())

    target = tmp_path / "failure" / "out.xlsx"

    with pytest.raises(RuntimeError):
        write_xlsx_atomic({"Sheet": pd.DataFrame({"v": [1]})}, target)

    assert not target.exists()
    assert list(target.parent.glob("*.xlsx")) == []
