"""تست یکپارچهٔ قالب‌بندی Excel در لایهٔ Infra."""

from __future__ import annotations

from pathlib import Path
import logging
from typing import List, Set, Tuple

import importlib

import pandas as pd
import pandas.testing as tm
import pytest
from openpyxl import load_workbook
from pandas.api.types import is_datetime64_any_dtype

from app.infra.excel import exporter as excel_exporter  # noqa: E402
from app.infra.io_utils import write_xlsx_atomic  # noqa: E402


def _available_engines() -> List[str]:
    engines: List[str] = []
    for name in ("openpyxl", "xlsxwriter"):
        if importlib.util.find_spec(name) is not None:
            engines.append(name)
    return engines


_ENGINES = _available_engines()
if not _ENGINES:
    pytest.skip("هیچ engine اکسل نصب نشده است", allow_module_level=True)


@pytest.mark.parametrize("engine", _ENGINES)
def test_vazir_font_size_enforced(tmp_path: Path, monkeypatch: pytest.MonkeyPatch, engine: str) -> None:
    df = pd.DataFrame({"ستون": [1, 2, 3], "نام": ["الف", "ب", "ج"]})
    out = tmp_path / f"{engine}-vazir.xlsx"
    monkeypatch.setenv("EXCEL_ENGINE", engine)

    write_xlsx_atomic({"گزارش": df}, out, font_name="Vazirmatn")

    print(f"excel size[{engine}]={out.stat().st_size}")

    wb = load_workbook(out)
    ws = wb[wb.sheetnames[0]]

    assert ws.freeze_panes == "A2"
    assert ws.cell(1, 1).font.name == "Vazirmatn"
    assert ws.cell(1, 1).font.size == 8
    if engine == "openpyxl":
        assert len(wb.named_styles) < 50


@pytest.mark.parametrize("engine", _ENGINES)
def test_non_vazir_font_uses_policy_size(tmp_path: Path, monkeypatch: pytest.MonkeyPatch, engine: str) -> None:
    df = pd.DataFrame({"val": [1, 2]})
    out = tmp_path / f"{engine}-tahoma.xlsx"
    monkeypatch.setenv("EXCEL_ENGINE", engine)

    write_xlsx_atomic({"Sheet": df}, out, font_name="Tahoma")

    wb = load_workbook(out)
    ws = wb[wb.sheetnames[0]]
    assert ws.cell(1, 1).font.name == "Tahoma"
    assert ws.cell(1, 1).font.size == 8


@pytest.mark.parametrize("engine", _ENGINES)
def test_table_and_autofilter_created(tmp_path: Path, monkeypatch: pytest.MonkeyPatch, engine: str) -> None:
    df = pd.DataFrame({"A": [1, 2, 3], "B": ["x", "y", "z"]})
    out = tmp_path / f"{engine}-table.xlsx"
    monkeypatch.setenv("EXCEL_ENGINE", engine)

    write_xlsx_atomic({"Sheet": df}, out, font_name="Tahoma")

    wb = load_workbook(out)
    ws = wb["Sheet"]

    tables = list(ws.tables.values())
    assert len(tables) == 1
    table = tables[0]
    assert table.displayName.startswith("tbl_")
    assert table.tableStyleInfo.name == "TableStyleLight1"
    assert table.ref == "A1:B4"


@pytest.mark.parametrize("engine", _ENGINES)
def test_style_caches_remain_minimal(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, engine: str
) -> None:
    df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
    out = tmp_path / f"{engine}-style-count.xlsx"
    monkeypatch.setenv("EXCEL_ENGINE", engine)

    created_formats: Set[Tuple[str | None, int | None, bool]] = set()
    created_styles: Set[str] = set()

    if engine == "xlsxwriter":
        original = excel_exporter.ensure_xlsxwriter_format

        def tracker(workbook, font, *, header=False):
            created_formats.add((font.name, font.size, header))
            return original(workbook, font, header=header)

        monkeypatch.setattr(excel_exporter, "ensure_xlsxwriter_format", tracker)
    else:
        original = excel_exporter.ensure_openpyxl_named_style

        def tracker(workbook, font):
            style_name = original(workbook, font)
            created_styles.add(style_name)
            return style_name

        monkeypatch.setattr(excel_exporter, "ensure_openpyxl_named_style", tracker)

    write_xlsx_atomic({"Sheet": df}, out, font_name="Tahoma")

    if engine == "xlsxwriter":
        assert len(created_formats) <= 3
    else:
        assert len(created_styles) == 1


@pytest.mark.parametrize("engine", _ENGINES)
def test_table_name_start_letter(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, engine: str
) -> None:
    df = pd.DataFrame({"val": [1, 2]})
    sheets = {"123": df, "__mix": df}
    out = tmp_path / f"{engine}-table-letter.xlsx"
    monkeypatch.setenv("EXCEL_ENGINE", engine)

    write_xlsx_atomic(sheets, out, font_name="Tahoma")

    wb = load_workbook(out)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        tables = list(ws.tables.values())
        if not tables:
            continue
        name = tables[0].displayName
        assert name.startswith("tbl_")
        slug = name[len("tbl_") :]
        assert slug and slug[0].isalpha()
        if sheet.isdigit():
            assert name.startswith("tbl_t_")


@pytest.mark.parametrize("engine", _ENGINES)
def test_table_headers_dedup(tmp_path: Path, monkeypatch: pytest.MonkeyPatch, engine: str) -> None:
    df = pd.DataFrame([[1, 2, 3]], columns=["Dup", "Dup", "Dup"])
    out = tmp_path / f"{engine}-dedup.xlsx"
    monkeypatch.setenv("EXCEL_ENGINE", engine)
    monkeypatch.setattr(
        "app.infra.io_utils._prepare_dataframe_for_excel", lambda frame: frame.copy()
    )

    write_xlsx_atomic({"Sheet": df}, out, font_name="Tahoma")

    wb = load_workbook(out)
    ws = wb["Sheet"]
    row = next(ws.iter_rows(min_row=1, max_row=1, max_col=3))
    headers = [cell.value for cell in row]
    assert headers == ["Dup", "Dup_2", "Dup_3"]
    tables = list(ws.tables.values())
    assert tables, "table must exist for dedup test"
    table = tables[0]
    table_columns = table.tableColumns
    if hasattr(table_columns, "tableColumn"):
        iterable = table_columns.tableColumn
    else:
        iterable = table_columns
    column_names = [col.name for col in iterable]
    assert column_names == ["Dup", "Dup_2", "Dup_3"]


@pytest.mark.parametrize("engine", _ENGINES)
def test_datetime_format_preserved_xlsxwriter(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, engine: str
) -> None:
    if engine != "xlsxwriter":
        pytest.skip("این تست مخصوص xlsxwriter است")

    df = pd.DataFrame(
        {
            "ts": pd.date_range("2024-01-01", periods=3, freq="D"),
            "value": [1, 2, 3],
        }
    )
    out = tmp_path / "datetime.xlsx"
    monkeypatch.setenv("EXCEL_ENGINE", engine)

    write_xlsx_atomic({"Sheet": df}, out, font_name="Tahoma")

    round_trip = pd.read_excel(out, engine="openpyxl")
    assert is_datetime64_any_dtype(round_trip["ts"])
    tm.assert_series_equal(
        round_trip["ts"].astype("datetime64[ns]"),
        df["ts"].astype("datetime64[ns]"),
        check_names=False,
    )


@pytest.mark.parametrize("engine", _ENGINES)
def test_font_warning_emitted_once(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, caplog: pytest.LogCaptureFixture, engine: str
) -> None:
    excel_exporter._FONT_WARNING_EMITTED = False  # type: ignore[attr-defined]

    df = pd.DataFrame({"A": [1]})
    out_first = tmp_path / f"{engine}-warn-1.xlsx"
    out_second = tmp_path / f"{engine}-warn-2.xlsx"
    monkeypatch.setenv("EXCEL_ENGINE", engine)

    with caplog.at_level(logging.WARNING):
        write_xlsx_atomic({"Sheet": df}, out_first, font_name="Tahoma")
        write_xlsx_atomic({"Sheet": df}, out_second, font_name="Tahoma")

    messages = [record.getMessage() for record in caplog.records if "جاسازی" in record.getMessage()]
    assert len(messages) == 1
