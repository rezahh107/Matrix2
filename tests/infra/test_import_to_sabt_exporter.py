from __future__ import annotations

from datetime import datetime
from pathlib import Path
import sys

import pandas as pd
import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from app.infra.excel.import_to_sabt import (  # noqa: E402
    apply_alias_rule,
    build_errors_frame,
    build_optional_sheet_frame,
    build_sheet2_frame,
    build_summary_frame,
    ensure_template_workbook,
    ImportToSabtExportError,
    load_exporter_config,
    prepare_allocation_export_frame,
    write_import_to_sabt_excel,
    _coalesce_duplicate_identifier_rows,
    _safe_merge,
)


def _sample_alloc_frame() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "student_id": "STD-1",
                "student_GF_FirstName": "سارا",
                "student_GF_LastName": "محمدی",
                "student_GF_Mobile": "۹۱۲۳۴۵۶۷۸۹",
                "student_GF_NationalCode": "۰۰۱۲۳۴۵۶۷۸",
                "student_gender": 0,
                "student_graduation_status": 0,
                "student_center": 1,
                "student_finance": 3,
                "student_group_code": 1201,
                "student_exam_group": "ریاضی",
                "student_school_code": 10101,
                "student_school_name": "دبیرستان بعثت",
                "student_school_code_1": 10101,
                "student_school_name_1": "دبیرستان بعثت",
                "student_school_code_2": pd.NA,
                "student_school_name_2": pd.NA,
                "student_school_code_3": pd.NA,
                "student_school_name_3": pd.NA,
                "student_school_code_4": pd.NA,
                "student_school_name_4": pd.NA,
                "registration_status": "3",
                "hekmat_tracking": "۱۲۳۴۵۶۷۸۹۰۱۲۳۴۵۶",
                "hekmat_package": "طلایی",
                "mentor_postal_code": "۱۲۳۴۵",
                "mentor_alias_postal_code": "54321",
                "mentor_mentor_name": "پشتیبان حکمت",
                "mentor_mentor_id": "EMP-1",
                "mentor_manager_name": "مدیر حکمت",
                "allocation_status": "success",
            },
            {
                "student_id": "STD-2",
                "student_GF_FirstName": "علی",
                "student_GF_LastName": "حسینی",
                "student_GF_Mobile": "09123456789",
                "student_GF_NationalCode": "1234567890",
                "student_gender": 1,
                "student_graduation_status": 1,
                "student_center": 2,
                "student_finance": 0,
                "student_group_code": 1502,
                "student_exam_group": "تجربی",
                "student_school_code": 20202,
                "student_school_name": "دبیرستان اندیشه",
                "student_school_code_1": 20202,
                "student_school_name_1": "دبیرستان اندیشه",
                "student_school_code_2": pd.NA,
                "student_school_name_2": pd.NA,
                "student_school_code_3": pd.NA,
                "student_school_name_3": pd.NA,
                "student_school_code_4": pd.NA,
                "student_school_name_4": pd.NA,
                "registration_status": "0",
                "hekmat_tracking": "654321",
                "hekmat_package": "نقره‌ای",
                "mentor_postal_code": "67890",
                "mentor_alias_postal_code": "",
                "mentor_mentor_name": "پشتیبان عادی",
                "mentor_mentor_id": "EMP-2",
                "mentor_manager_name": "مدیر عادی",
                "allocation_status": "failed",
            },
        ]
    )


def test_sheet2_structure_matches_template(tmp_path: Path) -> None:
    cfg = load_exporter_config("config/SmartAlloc_Exporter_Config_v1.json")
    df_alloc = _sample_alloc_frame()
    df_sheet2 = build_sheet2_frame(df_alloc, cfg, today=datetime(2024, 3, 20))
    df_sheet2 = apply_alias_rule(df_sheet2, df_alloc)
    logs_df = pd.DataFrame(
        [
            {"student_id": "STD-1", "allocation_status": "success"},
            {
                "student_id": "STD-2",
                "allocation_status": "failed",
                "error_type": "CAPACITY",
                "detailed_reason": "Full",
            },
        ]
    )
    df_summary = build_summary_frame(
        cfg,
        total_students=5,
        allocated_count=len(df_sheet2),
        error_count=1,
    )
    df_errors = build_errors_frame(logs_df, cfg)
    df_sheet5 = build_optional_sheet_frame(cfg, "Sheet5")
    df_9394 = build_optional_sheet_frame(cfg, "9394")
    template = tmp_path / "template.xlsx"
    ensure_template_workbook(template, cfg)
    output = tmp_path / "sabt.xlsx"
    write_import_to_sabt_excel(
        df_sheet2,
        df_summary,
        df_errors,
        df_sheet5,
        df_9394,
        template,
        output,
    )
    assert output.exists()
    df_result = pd.read_excel(output, sheet_name="Sheet2")
    df_template = pd.read_excel(template, sheet_name="Sheet2")
    assert df_result.columns.tolist() == df_template.columns.tolist()
    assert len(df_result.columns) == 46
    assert df_result.loc[0, "شناسه دانش آموز"] == "STD-1"
    assert str(df_result.iloc[0]["کد ملی"]).zfill(10) == "0012345678"
    assert str(df_result.iloc[0]["تلفن همراه"]).zfill(11) == "09123456789"
    assert str(df_result.iloc[0]["کد پستی"]).zfill(10) == "0000054321"


def test_write_import_to_sabt_excel_repairs_mismatched_headers(tmp_path: Path) -> None:
    cfg = load_exporter_config("config/SmartAlloc_Exporter_Config_v1.json")
    df_alloc = _sample_alloc_frame()
    df_sheet2 = build_sheet2_frame(df_alloc, cfg, today=datetime(2024, 3, 20))
    df_sheet2 = apply_alias_rule(df_sheet2, df_alloc)
    df_summary = build_summary_frame(
        cfg,
        total_students=2,
        allocated_count=1,
        error_count=1,
    )
    df_errors = build_errors_frame(
        pd.DataFrame(
            [
                {"student_id": "STD-1", "allocation_status": "success"},
                {"student_id": "STD-2", "allocation_status": "failed"},
            ]
        ),
        cfg,
    )
    df_sheet5 = build_optional_sheet_frame(cfg, "Sheet5")
    df_9394 = build_optional_sheet_frame(cfg, "9394")

    workbook = Workbook()
    ws_sheet2 = workbook.active
    ws_sheet2.title = "Sheet2"
    sheet2_expected = len(cfg["sheets"]["Sheet2"]["columns"])
    for idx in range(sheet2_expected + 3):
        ws_sheet2.cell(row=1, column=idx + 1, value=f"قدیمی {idx + 1}")

    summary = workbook.create_sheet("Summary")
    summary_expected = len(cfg["sheets"]["Summary"]["columns"])
    for idx in range(summary_expected):
        summary.cell(row=1, column=idx + 1, value=f"سابق {idx + 1}")

    errors = workbook.create_sheet("Errors")
    errors_expected = len(cfg["sheets"]["Errors"]["columns"])
    for idx in range(errors_expected + 1):
        errors.cell(row=1, column=idx + 1, value=f"ستون {idx + 1}")

    sheet5 = workbook.create_sheet("Sheet5")
    sheet5_expected = len(cfg["sheets"]["Sheet5"]["columns"])
    for idx in range(sheet5_expected):
        sheet5.cell(row=1, column=idx + 1, value=f"۵-{idx + 1}")

    sheet9394 = workbook.create_sheet("9394")
    sheet9394_expected = len(cfg["sheets"]["9394"]["columns"])
    for idx in range(sheet9394_expected):
        sheet9394.cell(row=1, column=idx + 1, value=f"۹۳-{idx + 1}")
    template = tmp_path / "template.xlsx"
    workbook.save(template)

    output = tmp_path / "sabt.xlsx"
    write_import_to_sabt_excel(
        df_sheet2,
        df_summary,
        df_errors,
        df_sheet5,
        df_9394,
        template,
        output,
    )

    expected_columns = list(cfg["sheets"]["Sheet2"]["columns"].keys())
    df_result = pd.read_excel(output, sheet_name="Sheet2")
    assert df_result.columns.tolist() == expected_columns


def test_ensure_template_backfills_missing_summary(tmp_path: Path) -> None:
    cfg = load_exporter_config("config/SmartAlloc_Exporter_Config_v1.json")
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    ws_sheet2 = workbook.active
    ws_sheet2.title = "Sheet2"
    workbook.save(template)

    ensure_template_workbook(template, cfg)

    from openpyxl import load_workbook

    loaded = load_workbook(template)
    assert "Summary" in loaded.sheetnames
    summary = loaded["Summary"]
    expected_headers = cfg["sheets"]["Summary"]["columns"]
    header_values = [cell.value for cell in next(summary.iter_rows(min_row=1, max_row=1))]
    assert header_values[: len(expected_headers)] == expected_headers


def test_hekmat_rule_resets_non_matching_rows() -> None:
    cfg = load_exporter_config("config/SmartAlloc_Exporter_Config_v1.json")
    df_alloc = _sample_alloc_frame()
    df_sheet2 = build_sheet2_frame(df_alloc, cfg)
    assert df_sheet2.loc[0, "کد رهگیری حکمت"] == "1234567890123456"
    assert df_sheet2.loc[1, "کد رهگیری حکمت"] == ""
    assert df_sheet2.loc[1, "نوع بسته حکمت"] == ""


def test_mobile_normalizer_handles_various_inputs() -> None:
    cfg = load_exporter_config("config/SmartAlloc_Exporter_Config_v1.json")
    df_alloc = _sample_alloc_frame().copy()
    df_alloc.loc[0, "student_GF_Mobile"] = "۹۱۲۳۴۵۶۷۸۹"
    df_alloc.loc[1, "student_GF_Mobile"] = "9123456789"
    df_sheet2 = build_sheet2_frame(df_alloc, cfg)
    assert df_sheet2.loc[0, "تلفن همراه"] == "09123456789"
    assert df_sheet2.loc[1, "تلفن همراه"] == "09123456789"


def test_alias_rule_prefers_alias_when_present() -> None:
    cfg = load_exporter_config("config/SmartAlloc_Exporter_Config_v1.json")
    df_alloc = _sample_alloc_frame()
    df_sheet2 = build_sheet2_frame(df_alloc, cfg)
    df_sheet2 = apply_alias_rule(df_sheet2, df_alloc)
    assert df_sheet2.loc[0, "کد پستی"] == "0000054321"
    assert df_sheet2.loc[1, "کد پستی"] == "0000067890"


def test_prepare_allocation_export_frame_preserves_length_and_index() -> None:
    alloc = pd.DataFrame(
        [
            {"student_id": "STD-1", "mentor_id": "EMP-1", "allocation_status": "success"},
            {"student_id": "STD-2", "mentor_id": "EMP-2", "allocation_status": "success"},
        ]
    )
    students = pd.DataFrame(
        [
            {"student_id": "STD-1", "GF_FirstName": "سارا"},
            {"student_id": "STD-2", "GF_FirstName": "علی"},
        ]
    )
    mentors = pd.DataFrame(
        [
            {"mentor_id": "EMP-1", "mentor_name": "پشتیبان ۱"},
            {"mentor_id": "EMP-2", "mentor_name": "پشتیبان ۲"},
        ]
    )

    merged = prepare_allocation_export_frame(alloc, students, mentors)

    assert len(merged) == len(alloc)
    assert list(merged.index) == list(alloc.index)


def test_prepare_allocation_export_frame_reports_missing_student_identifier(tmp_path: Path) -> None:
    alloc = pd.DataFrame([
        {"student_id": "STD-1", "mentor_id": "EMP-1"},
    ])
    students = pd.DataFrame([
        {"GF_FirstName": "سارا"},
    ])
    students.attrs["sheet_name"] = "Students"
    students.attrs["source_path"] = tmp_path / "students.xlsx"
    mentors = pd.DataFrame([
        {"mentor_id": "EMP-1", "mentor_name": "پشتیبان"},
    ])

    with pytest.raises(ImportToSabtExportError) as excinfo:
        prepare_allocation_export_frame(alloc, students, mentors)

    message = str(excinfo.value)
    assert "ImportToSabt export failed: missing required student identifier" in message
    assert "sheet 'Students'" in message
    assert "students.xlsx" in message
    assert "Please add the column" in message


def test_prepare_allocation_export_frame_reports_missing_mentor_identifier(tmp_path: Path) -> None:
    alloc = pd.DataFrame([
        {"student_id": "STD-1", "mentor_id": "EMP-1"},
    ])
    students = pd.DataFrame([
        {"student_id": "STD-1", "GF_FirstName": "سارا"},
    ])
    mentors = pd.DataFrame([
        {"mentor_name": "پشتیبان"},
    ])
    mentors.attrs["sheet_name"] = "Mentors"
    mentors.attrs["source_path"] = tmp_path / "mentors.xlsx"

    with pytest.raises(ImportToSabtExportError) as excinfo:
        prepare_allocation_export_frame(alloc, students, mentors)

    message = str(excinfo.value)
    assert "ImportToSabt export failed: missing required mentor identifier" in message
    assert "sheet 'Mentors'" in message
    assert "mentors.xlsx" in message
    assert "'mentor_id' یا 'alias'" in message
    assert "Please add the column" in message


def test_prepare_allocation_export_frame_rejects_duplicate_students() -> None:
    alloc = pd.DataFrame([
        {"student_id": "STD-1", "mentor_id": "EMP-1"},
    ])
    students = pd.DataFrame(
        [
            {"student_id": "STD-1", "GF_FirstName": "سارا"},
            {"student_id": "STD-1", "GF_FirstName": "زهرا"},
        ]
    )
    mentors = pd.DataFrame([
        {"mentor_id": "EMP-1", "mentor_name": "پشتیبان ۱"},
    ])

    with pytest.raises(ImportToSabtExportError) as excinfo:
        prepare_allocation_export_frame(alloc, students, mentors)

    message = str(excinfo.value)
    assert "duplicate" in message
    assert "STD-1" in message


def test_safe_merge_duplicate_error_includes_samples() -> None:
    left = pd.DataFrame({"student_id": ["STD-1", "STD-2"]})
    right = pd.DataFrame({"student_id": ["STD-1", "STD-1"]})

    with pytest.raises(ImportToSabtExportError) as excinfo:
        _safe_merge(
            left,
            right,
            context="student",
            on="student_id",
            validate="many_to_one",
            left_label="allocations",
            right_label="students",
        )

    message = str(excinfo.value)
    assert "duplicate keys" in message
    assert "allocations" in message
    assert "students" in message
    assert "STD-1" in message


def test_prepare_allocation_export_frame_coalesces_duplicate_mentor_rows() -> None:
    alloc = pd.DataFrame([
        {"student_id": "STD-1", "mentor_id": "EMP-1"},
    ])
    students = pd.DataFrame([
        {"student_id": "STD-1", "GF_FirstName": "سارا"},
    ])
    mentors = pd.DataFrame(
        [
            {
                "mentor_id": "EMP-1",
                "mentor_name": "",
                "mentor_postal_code": "",
            },
            {
                "mentor_id": "EMP-1",
                "mentor_name": "پشتیبان نهایی",
                "mentor_postal_code": "54321",
            },
        ]
    )

    merged = prepare_allocation_export_frame(alloc, students, mentors)

    assert merged.loc[0, "mentor_mentor_name"] == "پشتیبان نهایی"
    assert merged.loc[0, "mentor_mentor_postal_code"] == "54321"


def test_prepare_allocation_export_frame_handles_duplicate_columns_after_canon() -> None:
    alloc = pd.DataFrame(
        [
            {
                "student_id": "STD-1",
                "mentor_id": "EMP-1",
                "کد کارمندی پشتیبان": "EMP-1",
            }
        ]
    )
    students = pd.DataFrame(
        [
            {
                "student_id": "STD-1",
                "GF_FirstName": "سارا",
                "کد ملی": "0012345678",
            }
        ]
    )
    mentors = pd.DataFrame(
        [
            {
                "mentor_id": "EMP-1",
                "کد کارمندی پشتیبان": "EMP-1",
                "mentor_name": "پشتیبان الف",
            }
        ]
    )

    merged = prepare_allocation_export_frame(alloc, students, mentors)

    assert merged.loc[0, "student_GF_FirstName"] == "سارا"
    assert merged.loc[0, "mentor_mentor_name"] == "پشتیبان الف"


def test_prepare_allocation_export_frame_handles_duplicate_numeric_columns() -> None:
    alloc = pd.DataFrame(
        [
            {"student_id": "STD-1", "mentor_id": "EMP-1"},
        ]
    )
    students = pd.DataFrame(
        [
            {
                "student_id": "STD-1",
                "school_code": 101,
                "کد مدرسه": "00101",
            }
        ]
    )
    mentors = pd.DataFrame(
        [
            {
                "mentor_id": "EMP-1",
                "school_code": 202,
                "کد مدرسه": "00202",
            }
        ]
    )

    merged = prepare_allocation_export_frame(alloc, students, mentors)

    assert merged.loc[0, "student_school_code"] == 101
    assert merged.loc[0, "mentor_school_code"] == 202


def test_prepare_allocation_export_frame_raises_on_conflicting_duplicate_columns() -> None:
    alloc = pd.DataFrame(
        [
            {"student_id": "STD-1", "mentor_id": "EMP-1"},
        ]
    )
    students = pd.DataFrame(
        [
            {
                "student_id": "STD-1",
                "GF_FirstName": "سارا",
                "کد ملی": "001",
                "national_id": "999",
            }
        ]
    )
    mentors = pd.DataFrame(
        [
            {"mentor_id": "EMP-1", "mentor_name": "پشتیبان ۱"},
        ]
    )

    with pytest.raises(ImportToSabtExportError) as excinfo:
        prepare_allocation_export_frame(alloc, students, mentors)

    assert "duplicate columns" in str(excinfo.value)

def test_coalesce_duplicate_identifier_rows_handles_duplicate_columns() -> None:
    frame = pd.DataFrame(
        {
            "student_id": ["A1", "A1"],
            "value": [pd.NA, "filled"],
        }
    )
    frame.insert(0, "dup_id", frame["student_id"])
    frame.columns = ["student_id", "student_id", "value"]

    result = _coalesce_duplicate_identifier_rows(
        frame,
        "student_id",
        entity_name="student",
    )

    assert result.shape[0] == 1
    assert result.loc[result.index[0], "value"] == "filled"


def test_coalesce_duplicate_identifier_rows_handles_non_range_index() -> None:
    frame = pd.DataFrame(
        {
            "student_id": ["B1", "B1"],
            "value": ["first", "second"],
        },
        index=[10, 50],
    )

    result = _coalesce_duplicate_identifier_rows(
        frame,
        "student_id",
        entity_name="student",
    )

    assert result.shape[0] == 1
    assert list(result["value"]) == ["first"]
