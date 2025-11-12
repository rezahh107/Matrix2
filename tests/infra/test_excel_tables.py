"""صحت‌سنجی نام‌گذاری و رفتار جدول‌های Excel."""

from __future__ import annotations

from pathlib import Path
import sys
from typing import List

import importlib

import pandas as pd
import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

from app.infra.excel.tables import TableNameRegistry  # noqa: E402
from app.infra.io_utils import write_xlsx_atomic  # noqa: E402


def _engines() -> List[str]:
    engines: List[str] = []
    for name in ("openpyxl", "xlsxwriter"):
        if importlib.util.find_spec(name) is not None:
            engines.append(name)
    return engines


_ENGINES = _engines()
if not _ENGINES:
    pytest.skip("هیچ engine اکسل موجود نیست", allow_module_level=True)


@pytest.mark.parametrize("engine", _ENGINES)
def test_table_names_are_unique_and_slugged(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, engine: str
) -> None:
    data = pd.DataFrame({"الف": [1, 2], "ب": [3, 4]})
    sheets = {
        " 1/مالی ": data,
        "1/مالی": data,
        "مدرسه": data,
    }
    out = tmp_path / f"{engine}-slug.xlsx"
    monkeypatch.setenv("EXCEL_ENGINE", engine)

    write_xlsx_atomic(sheets, out, font_name="Vazirmatn")

    wb = load_workbook(out)
    seen: set[str] = set()
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        tables = list(ws.tables.values())
        if not tables:
            continue
        assert len(tables) == 1
        table = tables[0]
        assert table.displayName.startswith("tbl_")
        assert table.displayName not in seen
        seen.add(table.displayName)
    assert len(seen) == len(sheets)


@pytest.mark.parametrize("engine", _ENGINES)
def test_table_not_created_for_empty_frames(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, engine: str
) -> None:
    df = pd.DataFrame({"الف": pd.Series(dtype=int)})
    out = tmp_path / f"{engine}-empty.xlsx"
    monkeypatch.setenv("EXCEL_ENGINE", engine)

    write_xlsx_atomic({"خالی": df}, out, font_name="Vazirmatn")

    wb = load_workbook(out)
    ws = wb[wb.sheetnames[0]]
    assert not list(ws.tables.values())


def test_registry_instances_are_isolated() -> None:
    first = TableNameRegistry()
    second = TableNameRegistry()

    assert first.reserve("Sheet") == "tbl_Sheet"
    assert second.reserve("Sheet") == "tbl_Sheet"
