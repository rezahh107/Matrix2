"""تست‌های جامع برای import_to_sabt با داده‌های واقعی."""

from __future__ import annotations

from collections import OrderedDict
from pathlib import Path

import pandas as pd
import pytest

from app.infra.excel.import_to_sabt import (
    GF_FIELD_TO_COL,
    apply_alias_rule,
    build_errors_frame,
    build_sheet2_frame,
    build_summary_frame,
    load_exporter_config,
    prepare_allocation_export_frame,
    write_import_to_sabt_excel,
    _apply_normalizers,
    _normalize_mobile_ir,
)


@pytest.fixture
def config_path() -> Path:
    return Path("docs/SmartAlloc_Exporter_Config_v1.json")


@pytest.fixture
def exporter_config(config_path: Path) -> dict:
    return load_exporter_config(config_path)


@pytest.fixture
def sample_allocation_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "student_id": ["S001", "S002", "S003"],
            "mentor_id": ["M001", "M002", "M003"],
            "student_first_name": ["علی", "فاطمه", "محمد"],
            "student_last_name": ["احمدی", "رضایی", "کریمی"],
            "student_father_name": ["حسن", "رضا", "علی"],
            "student_national_code": ["1234567890", "0987654321", "1122334455"],
            "student_gender": ["M", "F", "M"],
            "student_educational_status": ["دانش‌آموز", "دانش‌آموز", "فارغ‌التحصیل"],
            "student_mobile": ["09121234567", "09359876543", "09181112222"],
            "contact1_mobile": ["09121111111", "09352222222", "09183333333"],
            "contact2_mobile": ["09124444444", "", "09185555555"],
            "student_exam_group": ["تجربی", "ریاضی", "انسانی"],
            "student_school_code": ["1001", "1002", "9000"],
            "student_school_name": ["", "", "مدرسه شهید بهشتی"],
            "student_average": [18.5, 19.2, 17.8],
            "student_registration_status": ["0", "1", "3"],
            "student_center": ["1", "1", "2"],
            "student_hekmat_tracking_code": ["", "", "1234567890123456"],
            "student_hekmat_package_type": ["", "", "ویژه"],
            "student_class_number": ["12", "11", ""],
            "student_konkur_quota": ["منطقه1", "منطقه2", ""],
            "student_seat_number": ["101", "102", ""],
            "student_created_at": ["2024-01-15", "2024-01-16", "2024-01-17"],
            "mentor_alias_code": ["1234567890", "", ""],
            "mentor_is_school_limited": [False, True, True],
        }
    )


@pytest.fixture
def school_city_lookup() -> dict[str, str]:
    return {"1001": "تهران", "1002": "اصفهان", "9000": "سایر"}


class TestConfigLoading:
    def test_config_loads_with_order_preserved(self, exporter_config: dict) -> None:
        columns = exporter_config["sheets"]["Sheet2"]["columns"]
        assert isinstance(columns, (dict, OrderedDict))
        assert next(iter(columns.keys())) == "پشتیبان"
        assert len(columns) == 46

    def test_config_column_names_match_spec(self, exporter_config: dict, config_path: Path) -> None:
        columns = list(exporter_config["sheets"]["Sheet2"]["columns"].keys())
        spec_path = config_path.parent / "ImportToSabt1404_Sheet2_columns.md"
        with spec_path.open(encoding="utf-8") as handle:
            lines = handle.readlines()
        expected_columns: list[str] = []
        for line in lines:
            stripped = line.strip()
            if stripped and stripped[0].isdigit() and ". " in stripped:
                expected_columns.append(stripped.split(". ", 1)[1])
        assert columns == expected_columns


class TestGFFieldMapping:
    def test_all_critical_fields_mapped(self) -> None:
        for field_id in ["101", "102", "143", "92", "73", "75", "30", "20", "21"]:
            assert field_id in GF_FIELD_TO_COL

    def test_field_mapping_includes_persian_names(self) -> None:
        assert "نام" in GF_FIELD_TO_COL["101"]
        assert "نام خانوادگی" in GF_FIELD_TO_COL["102"]


class TestSheet2Generation:
    def test_sheet2_has_correct_column_count(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        assert sheet2.shape[1] == 46

    def test_sheet2_column_order_preserved(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        expected_order = list(exporter_config["sheets"]["Sheet2"]["columns"].keys())
        assert list(sheet2.columns) == expected_order

    def test_sheet2_row_count_matches_input(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        assert len(sheet2) == len(sample_allocation_df)

    def test_empty_columns_are_empty(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        assert sheet2["پشتیبان"].str.strip().eq("").all()
        assert sheet2["کد ثبت نام"].str.strip().eq("").all()

    def test_gender_mapping_applied(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        assert sheet2.loc[0, "جنسیت"] == "پسر"
        assert sheet2.loc[1, "جنسیت"] == "دختر"

    def test_registration_status_mapping(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        assert sheet2.loc[0, "وضعیت ثبت نام"] == "عادی"
        assert sheet2.loc[1, "وضعیت ثبت نام"] == "بنیاد"
        assert sheet2.loc[2, "وضعیت ثبت نام"] == "حکمت"

    def test_national_code_normalized(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        for value in sheet2["کد ملی"]:
            if value:
                assert len(value) == 10
                assert value.isdigit()

    def test_mobile_normalized(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        for value in sheet2["تلفن همراه"]:
            if value:
                assert len(value) == 11
                assert value.startswith("09")

    def test_average_precision(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        avg = sheet2.loc[0, "معدل"]
        if avg:
            assert len(str(float(avg)).split(".")[-1]) <= 2


class TestHekmatConditionalLogic:
    def test_hekmat_fields_only_for_status_3(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        assert sheet2.loc[0, "کد رهگیری حکمت"] == ""
        assert sheet2.loc[1, "کد رهگیری حکمت"] == ""
        assert sheet2.loc[2, "کد رهگیری حکمت"] == "1111111111111111"

    def test_hekmat_code_normalized(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        hekmat_code = sheet2.loc[2, "کد رهگیری حکمت"]
        if hekmat_code:
            assert len(hekmat_code) == 16
            assert hekmat_code.isdigit()


class TestRegistrationStatusPreservation:
    def test_prefixed_status_values_survive_mapping(
        self, exporter_config: dict
    ) -> None:
        allocations_df = pd.DataFrame({"student_id": ["s1", "s2", "s3"], "mentor_id": [1, 2, 3]})
        students_df = pd.DataFrame(
            {
                "student_id": ["s1", "s2", "s3"],
                "student_first_name": ["الف", "ب", "پ"],
                "student_last_name": ["یک", "دو", "سه"],
                "student_gender": ["M", "F", "M"],
                "وضعیت ثبت نام": [0, 3, pd.NA],
            }
        )
        mentors_df = pd.DataFrame({"mentor_id": [1, 2, 3]})

        prepared = prepare_allocation_export_frame(
            allocations_df, students_df, mentors_df, student_ids=allocations_df["student_id"]
        )
        sheet2 = build_sheet2_frame(prepared, exporter_config)

        assert sheet2.loc[0, "وضعیت ثبت نام"] == "عادی"
        assert sheet2.loc[1, "وضعیت ثبت نام"] == "حکمت"
        assert sheet2.loc[2, "وضعیت ثبت نام"] == "عادی"

    def test_cli_path_preserves_reg_status_column(
        self, exporter_config: dict
    ) -> None:
        allocations_df = pd.DataFrame(
            {
                "student_id": ["s1", "s2", "s3"],
                "mentor_id": [1, 2, 3],
            }
        )
        students_df = pd.DataFrame(
            {
                "student_id": ["s1", "s2", "s3"],
                "reg_status": [0, 3, 0],
                "student_first_name": ["الف", "ب", "پ"],
                "student_last_name": ["یک", "دو", "سه"],
                "student_gender": ["M", "F", "M"],
            }
        )
        mentors_df = pd.DataFrame({"mentor_id": [1, 2, 3]})

        prepared = prepare_allocation_export_frame(
            allocations_df,
            students_df,
            mentors_df,
            student_ids=allocations_df["student_id"],
        )

        assert "student_registration_status" in prepared.columns
        numeric_status = prepared["student_registration_status"].astype("Int64").tolist()
        assert numeric_status == [0, 3, 0]

        sheet2 = build_sheet2_frame(prepared, exporter_config)

        assert sheet2["وضعیت ثبت نام"].tolist() == ["عادی", "حکمت", "عادی"]

    def test_registration_status_persian_digits(self, exporter_config: dict) -> None:
        minimal_df = pd.DataFrame(
            {
                "student_id": ["s1", "s2"],
                "mentor_id": [1, 2],
                "student_first_name": ["الف", "ب"],
                "student_last_name": ["یک", "دو"],
                "student_gender": ["M", "F"],
                "student_registration_status": ["۰", "۳"],
            }
        )

        sheet2 = build_sheet2_frame(minimal_df, exporter_config)

        assert sheet2.loc[0, "وضعیت ثبت نام"] == "عادی"
        assert sheet2.loc[1, "وضعیت ثبت نام"] == "حکمت"

    def test_debug_log_traces_registration_status(
        self, exporter_config: dict
    ) -> None:
        allocations_df = pd.DataFrame({"student_id": ["s1", "s2"], "mentor_id": [1, 2]})
        students_df = pd.DataFrame(
            {
                "student_id": ["s1", "s2"],
                "student_first_name": ["الف", "ب"],
                "student_last_name": ["یک", "دو"],
                "student_gender": ["M", "F"],
                "وضعیت ثبت نام": [0, 3],
            }
        )
        mentors_df = pd.DataFrame({"mentor_id": [1, 2]})

        prepared = prepare_allocation_export_frame(
            allocations_df, students_df, mentors_df, student_ids=allocations_df["student_id"]
        )

        debug_log = prepared.attrs.get("registration_status_debug")
        assert isinstance(debug_log, list)
        debug_lookup = {entry.get("label"): entry for entry in debug_log}
        assert debug_lookup["students_raw"]["threes"] == 1
        assert debug_lookup["merged_after_enrich"]["zeros"] == 1

        sheet2 = build_sheet2_frame(prepared, exporter_config)
        sheet_debug = sheet2.attrs.get("registration_status_debug")
        assert isinstance(sheet_debug, list)
        sheet_lookup = {entry.get("label"): entry for entry in sheet_debug}
        assert sheet_lookup["sheet2_status_normalized"]["threes"] == 1


class TestDerivedFields:
    def test_city_derived_from_school_code(
        self,
        sample_allocation_df: pd.DataFrame,
        exporter_config: dict,
        school_city_lookup: dict[str, str],
    ) -> None:
        exporter_config.setdefault("lookups", {})["city_by_school_code"] = school_city_lookup
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        assert sheet2.loc[0, "شهر مدرسه 1"] == "تهران"
        assert sheet2.loc[1, "شهر مدرسه 1"] == "اصفهان"
        assert sheet2.loc[2, "شهر مدرسه 1"] == "سایر"

    def test_school_code_9000_uses_name(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        assert sheet2.loc[2, "کد مدرسه 1"] in {"9000", "مدرسه شهید بهشتی"}


class TestAliasRule:
    def test_postal_code_filled_from_mentor_alias(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        sheet2 = apply_alias_rule(sheet2, sample_allocation_df)
        assert sheet2.loc[0, "کد پستی"] == "1234567890"

    def test_alias_column_filled_for_school_limited(
        self, sample_allocation_df: pd.DataFrame, exporter_config: dict
    ) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        sheet2 = apply_alias_rule(sheet2, sample_allocation_df)
        if "کد پستی جایگزین" in sheet2.columns:
            assert sheet2.loc[1, "کد پستی جایگزین"] == "M002"
            assert sheet2.loc[2, "کد پستی جایگزین"] == "M003"

    def test_non_school_limited_has_empty_alias(
        self, sample_allocation_df: pd.DataFrame, exporter_config: dict
    ) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        sheet2 = apply_alias_rule(sheet2, sample_allocation_df)
        if "کد پستی جایگزین" in sheet2.columns:
            alias_val = sheet2.loc[0, "کد پستی جایگزین"]
            assert alias_val in {"", None}


class TestDateFormatting:
    def test_registration_date_format(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        date_str = sheet2.loc[0, "تاریخ ثبت نام"]
        if date_str:
            parts = date_str.split("-")
            assert len(parts) == 3
            assert len(parts[0]) == 4
            assert len(parts[1]) == 2
            assert len(parts[2]) == 2


class TestEdgeCases:
    def test_empty_contact2_mobile_stays_empty(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        assert sheet2.loc[1, "تلفن رابط 2"] == ""

    def test_missing_columns_result_in_empty_values(self, exporter_config: dict) -> None:
        minimal_df = pd.DataFrame({"student_id": ["S001"], "mentor_id": ["M001"], "student_first_name": ["علی"]})
        sheet2 = build_sheet2_frame(minimal_df, exporter_config)
        assert sheet2.loc[0, "نام پدر"] == ""
        assert sheet2.loc[0, "کد ملی"] == ""

    def test_unicode_and_special_chars_preserved(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        sample_allocation_df.loc[0, "student_first_name"] = "علی‌رضا"
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        assert "علی" in sheet2.loc[0, "نام"]


class TestExcelOutput:
    def test_excel_file_created_successfully(
        self, sample_allocation_df: pd.DataFrame, exporter_config: dict, tmp_path: Path
    ) -> None:
        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        summary = build_summary_frame(exporter_config, total_students=3, allocated_count=3, error_count=0)
        errors = build_errors_frame(None, exporter_config)
        output_path = tmp_path / "output.xlsx"
        template_path = tmp_path / "template.xlsx"
        write_import_to_sabt_excel(sheet2, summary, errors, None, None, template_path, output_path)
        assert output_path.exists()

    def test_excel_headers_match_config(
        self, sample_allocation_df: pd.DataFrame, exporter_config: dict, tmp_path: Path
    ) -> None:
        from openpyxl import load_workbook

        sheet2 = build_sheet2_frame(sample_allocation_df, exporter_config)
        summary = build_summary_frame(exporter_config, total_students=3, allocated_count=3, error_count=0)
        errors = build_errors_frame(None, exporter_config)
        output_path = tmp_path / "output.xlsx"
        template_path = tmp_path / "template.xlsx"
        write_import_to_sabt_excel(sheet2, summary, errors, None, None, template_path, output_path)
        workbook = load_workbook(output_path)
        headers = [cell.value for cell in next(workbook["Sheet2"].iter_rows(min_row=1, max_row=1))]
        expected_headers = list(exporter_config["sheets"]["Sheet2"]["columns"].keys())
        assert headers[: len(expected_headers)] == expected_headers


class TestNormalizers:
    def test_apply_normalizers_none_returns_original(self, sample_allocation_df: pd.DataFrame, exporter_config: dict) -> None:
        series = pd.Series(["abc", "def"])
        result = _apply_normalizers(series, None)
        pd.testing.assert_series_equal(result, series)

    def test_digits_10_normalizer(self) -> None:
        series = pd.Series(["۱۲۳۴۵۶۷۸۹۰", "123-456-7890"])
        result = _apply_normalizers(series, "digits_10")
        assert result.tolist() == ["1234567890", "1234567890"]

    def test_digits_16_normalizer(self) -> None:
        series = pd.Series(["1234-5678-9012-3456", None])
        result = _apply_normalizers(series, "digits_16")
        assert result.tolist() == ["1234567890123456", ""]

    def test_mobile_ir_normalizer_variants(self) -> None:
        values = ["09121234567", "+989121234567", "9121234567", None]
        result = _apply_normalizers(pd.Series(values), "mobile_ir")
        assert result.tolist() == ["09121234567", "", "09121234567", ""]

    def test_combined_normalizers(self) -> None:
        series = pd.Series(["+989121234567"])
        result = _apply_normalizers(series, ["digits_16", "mobile_ir"])
        assert result.iloc[0] == ""

    def test_unknown_normalizer_raises(self) -> None:
        with pytest.raises(ValueError):
            _apply_normalizers(pd.Series(["1"]), "unknown_norm")

    def test_normalize_mobile_ir_direct(self) -> None:
        assert _normalize_mobile_ir("+989121234567") == ""
        assert _normalize_mobile_ir("9121234567") == "09121234567"
        assert _normalize_mobile_ir(None) == ""


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
