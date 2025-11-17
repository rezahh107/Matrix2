from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from app.infra.io_utils import write_xlsx_atomic


def _read_sheet(path: Path) -> tuple[list[str], list[list[object]], list[list[str]]]:
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in sheet[1]]
    values: list[list[object]] = []
    types: list[list[str]] = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        values.append([cell.value for cell in row])
        types.append([cell.data_type for cell in row])
    return headers, values, types


def test_allocations_sabt_phone_columns_roundtrip(tmp_path: Path) -> None:
    raw = pd.DataFrame(
        {
            "تلفن همراه داوطلب": [9300798195.0, "09123456789"],
            "تلفن همراه پدر": [9906421042.0, "09999999999"],
            "تلفن همراه مادر": [9174438540.0, pd.NA],
            "تلفن منزل": [9300012345.0, "09330001234"],
            "کد رهگیری حکمت": [1111111111111111, None],
            "نمره": [1, 2],
        }
    )

    output_path = tmp_path / "allocations_sabt.xlsx"
    write_xlsx_atomic({"allocations_sabt": raw}, output_path, header_mode=None)

    headers, rows, types = _read_sheet(output_path)

    col_index = {name: idx for idx, name in enumerate(headers)}
    expected_first_row = {
        "تلفن همراه داوطلب": "09300798195",
        "تلفن همراه پدر": "09906421042",
        "تلفن همراه مادر": "09174438540",
        "تلفن منزل": "09300012345",
        "کد رهگیری حکمت": "1111111111111111",
        "نمره": 1,
    }

    first_row = rows[0]
    first_types = types[0]
    for column, expected in expected_first_row.items():
        idx = col_index[column]
        assert first_row[idx] == expected
        if column == "نمره":
            assert first_types[idx] == "n"
        else:
            assert first_types[idx] in {"s", "inlineStr"}

    # ensure second row also stored as text for phone columns
    second_row = rows[1]
    second_types = types[1]
    for column in (
        "تلفن همراه داوطلب",
        "تلفن همراه پدر",
        "تلفن همراه مادر",
        "تلفن منزل",
        "کد رهگیری حکمت",
    ):
        idx = col_index[column]
        assert second_types[idx] in {"s", "inlineStr"}

