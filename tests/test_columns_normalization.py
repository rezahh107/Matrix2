from __future__ import annotations

from pathlib import Path
import sys

import importlib

import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.core.common import columns
from app.core.allocate_students import _require_columns
from app.infra.io_utils import write_xlsx_atomic


class _DummyPolicy:
    def __init__(self, alias_map: dict[str, dict[str, str]] | None = None):
        self.column_aliases = alias_map or {}


@pytest.fixture(autouse=True)
def _patch_policy(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(columns, "get_policy", lambda: _DummyPolicy())


def test_resolve_and_coerce_produce_canonical_columns() -> None:
    report_df = pd.DataFrame(
        {
            "وضعیت‌تحصیلی": ["فارغ التحصیل"],
            "مرکز ثبت نام": [" 12 "],
            "group_code": ["4001"],
            "کد مدرسه": ["5001.0"],
            "کد کارمندی پشتیبان": [4001.0],
            "کدپستی": ["0098"],
        }
    )
    inspactor_df = pd.DataFrame(
        {
            "کد گروه آزمایشی": ["3002"],
            "کد مدرسه 1": ["6005.0"],
            "mentor_id": ["4011.0"],
            "کدپستی": ["0215"],
        }
    )
    school_df = pd.DataFrame({"school_code": ["7003"], "نام‌مدرسه": ["نمونه"]})

    report_coerced = columns.coerce_semantics(
        columns.resolve_aliases(report_df, "report"),
        "report",
    )
    inspactor_coerced = columns.coerce_semantics(
        columns.resolve_aliases(inspactor_df, "inspactor"),
        "inspactor",
    )
    school_coerced = columns.coerce_semantics(
        columns.resolve_aliases(school_df, "school"),
        "school",
    )

    grad_col = columns.CANON_EN_TO_FA["graduation_status"]
    center_col = columns.CANON_EN_TO_FA["center"]
    group_col = columns.CANON_EN_TO_FA["group_code"]
    mentor_col = columns.CANON_EN_TO_FA["mentor_id"]
    postal_col = columns.CANON_EN_TO_FA["postal_code"]
    school_col = columns.CANON_EN_TO_FA["school_code"]

    assert {grad_col, center_col, group_col, mentor_col, postal_col, school_col}.issubset(
        report_coerced.columns
    )
    assert str(report_coerced[grad_col].dtype) in {"Int64", "int64"}
    assert int(report_coerced.loc[0, grad_col]) == 1
    assert str(report_coerced[center_col].dtype) in {"Int64", "int64"}
    assert int(report_coerced.loc[0, center_col]) == 12
    assert str(report_coerced[group_col].dtype) in {"Int64", "int64"}
    assert int(report_coerced.loc[0, group_col]) == 4001
    assert report_coerced[mentor_col].dtype == "string"
    assert report_coerced[postal_col].dtype == "string"
    assert str(report_coerced[school_col].dtype) in {"Int64", "int64"}

    assert str(inspactor_coerced[group_col].dtype) in {"Int64", "int64"}
    assert inspactor_coerced[mentor_col].dtype == "string"
    assert inspactor_coerced[postal_col].dtype == "string"
    assert str(
        inspactor_coerced[columns.CANON_EN_TO_FA["school_code_1"]].dtype
    ) in {"Int64", "int64"}

    assert str(school_coerced[school_col].dtype) in {"Int64", "int64"}
    assert school_coerced[columns.CANON_EN_TO_FA["school_name"]].dtype.name in {
        "object",
        "string",
    }


def test_canonicalize_headers_bilingual() -> None:
    df = pd.DataFrame({"کدرشته": [101], "نام مدرسه": ["Sample"]})
    bilingual = columns.canonicalize_headers(df, header_mode="fa_en")

    assert list(bilingual.columns) == [
        "کدرشته | group_code",
        "نام مدرسه | school_name",
    ]


def test_missing_column_error_reports_synonyms(monkeypatch: pytest.MonkeyPatch) -> None:
    df = pd.DataFrame({"group_code": [1]})

    with pytest.raises(ValueError) as exc:
        _require_columns(df, [columns.CANON_EN_TO_FA["group_code"]], "report")

    message = str(exc.value)
    assert "Missing columns" in message
    assert "accepted synonyms" in message
    assert "group_code" in message


@pytest.mark.skipif(
    importlib.util.find_spec("openpyxl") is None,
    reason="openpyxl برای بررسی خروجی لازم است",
)
def test_write_xlsx_atomic_preserves_integer_codes(tmp_path: Path) -> None:
    df = pd.DataFrame(
        {
            columns.CANON_EN_TO_FA["group_code"]: [4001],
            columns.CANON_EN_TO_FA["mentor_id"]: ["4011"],
        }
    )
    out = tmp_path / "matrix.xlsx"

    write_xlsx_atomic({"matrix": df}, out, header_mode="fa_en")

    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb["matrix"]
    assert ws.cell(row=1, column=1).value == "کدرشته | group_code"
    assert ws.cell(row=2, column=1).value == 4001
