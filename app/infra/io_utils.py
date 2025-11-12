"""ابزارهای ورودی/خروجی Excel در لایهٔ زیرساخت.

این ماژول صرفاً مسئول تطبیق داده‌ها با محدودیت‌های Excel است و هیچ منطق
دامنه‌ای در آن قرار ندارد تا اصل جداسازی Core/Infra حفظ شود.
"""

from __future__ import annotations

import contextlib
import json
import os
import re
import tempfile
import warnings
from os import PathLike
from pathlib import Path
from typing import Dict, Iterator, List, Sequence

import pandas as pd

from app.core.common.columns import CANON_EN_TO_FA, HeaderMode, canonicalize_headers
from app.core.policy_loader import get_policy

__all__ = [
    "ALT_CODE_COLUMN",
    "write_xlsx_atomic",
    "read_excel_first_sheet",
    "read_crosswalk_workbook",
]


ALT_CODE_COLUMN = "کد جایگزین"
_INVALID_SHEET_CHARS = re.compile(r"[\\/*?:\[\]]")
_STRING_EXPORT_KEYS: Sequence[str] = ("alias", "mentor_id", "postal_code")
_INT_EXPORT_KEYS: Sequence[str] = ("group_code", "school_code")


def _safe_sheet_name(name: str, taken: set[str]) -> str:
    """اصلاح و یکتا‌سازی نام شیت مطابق محدودیت‌های Excel."""

    base = _INVALID_SHEET_CHARS.sub(" ", (name or "Sheet").strip()) or "Sheet"
    base = base[:31]
    candidate = base
    index = 2
    while candidate in taken or not candidate:
        suffix = f" ({index})"
        candidate = (base[: max(0, 31 - len(suffix))] + suffix).rstrip()
        index += 1
    taken.add(candidate)
    return candidate


def _flatten_columns(df: pd.DataFrame) -> pd.DataFrame:
    """تخت‌سازی MultiIndex ستون‌ها برای سازگاری با Excel.

    مثال::

        >>> import pandas as pd
        >>> frame = pd.DataFrame({("الف", "ب"): [1, 2]})
        >>> _flatten_columns(frame).columns.tolist()
        ['الف__ب']
    """

    if not isinstance(df.columns, pd.MultiIndex):
        columns = [str(col).strip() if str(col).strip() else "column" for col in df.columns]
        return df.rename(columns=dict(zip(df.columns, columns)))

    flattened = ["__".join(map(str, level)).strip() for level in df.columns.to_flat_index()]
    cleaned = [col if col else "column" for col in flattened]
    return df.copy().set_axis(cleaned, axis="columns")


def _make_unique_columns(columns: Sequence[str]) -> List[str]:
    """ساخت نام ستون یکتا با حفظ ترتیب اولیه."""

    seen: dict[str, int] = {}
    unique: List[str] = []
    for original in columns:
        base = (original or "column").strip() or "column"
        count = seen.get(base, 0)
        suffix = "" if count == 0 else f" ({count + 1})"
        candidate = f"{base}{suffix}"
        while candidate in seen:
            count += 1
            suffix = f" ({count + 1})"
            candidate = f"{base}{suffix}"
        unique.append(candidate)
        seen[candidate] = 1
        seen[base] = count + 1
    return unique


def _coalesce_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """ادغام ستون‌های هم‌نام با پر کردن مقادیر خالی از راست به چپ."""

    if df.empty or df.columns.is_unique:
        return df.copy()

    labels = [str(col) for col in df.columns]
    groups: dict[str, List[int]] = {}
    for idx, label in enumerate(labels):
        groups.setdefault(label, []).append(idx)

    result = pd.DataFrame(index=df.index)
    for label, positions in groups.items():
        subset = df.iloc[:, positions]
        if isinstance(subset, pd.Series):
            result[label] = subset
        else:
            filled = subset.bfill(axis=1)
            result[label] = filled.iloc[:, 0]
    return result


def _stringify_cell(value: object) -> str:
    """تبدیل امن مقادیر پیچیده به رشته برای Excel."""

    if value is None:
        return ""
    try:
        if pd.isna(value):  # type: ignore[arg-type]
            return ""
    except Exception:
        pass
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8")
        except Exception:  # pragma: no cover - defensive branch
            return value.decode("latin-1", "ignore")
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, (pd.Series, pd.DataFrame)):
        return json.dumps(value.to_dict(), ensure_ascii=False, sort_keys=True)
    return str(value)


def _prepare_dataframe_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """پاک‌سازی کامل DataFrame پیش از نوشتن در Excel."""

    frame = _flatten_columns(df)
    frame = _coalesce_duplicate_columns(frame)
    frame.columns = _make_unique_columns(list(map(str, frame.columns)))

    converted = frame.copy()
    for column, dtype in converted.dtypes.items():
        if pd.api.types.is_object_dtype(dtype):
            converted[column] = converted[column].map(_stringify_cell)

    for key in _STRING_EXPORT_KEYS:
        fa_name = CANON_EN_TO_FA.get(key, key)
        for column_name in (fa_name, key):
            if column_name in converted.columns:
                converted[column_name] = converted[column_name].astype("string")

    for key in _INT_EXPORT_KEYS:
        fa_name = CANON_EN_TO_FA.get(key, key)
        for column_name in (fa_name, key):
            if column_name in converted.columns:
                numeric = pd.to_numeric(converted[column_name], errors="coerce")
                converted[column_name] = numeric.astype("Int64")

    return converted


def _build_table_name(sheet_name: str, taken: set[str]) -> str:
    """ساخت نام معتبر و یکتای جدول Excel بر اساس نام شیت."""

    base = re.sub(r"[^0-9A-Za-z_]", "_", (sheet_name or "Table").strip()) or "Table"
    if not re.match(r"[A-Za-z_]", base):
        base = f"T_{base}"
    base = base[:31]
    candidate = base
    index = 2
    while candidate in taken or not candidate:
        suffix = f"_{index}"
        candidate = (base[: max(0, 31 - len(suffix))] + suffix).rstrip("_")
        index += 1
    taken.add(candidate)
    return candidate


def _apply_excel_tables(
    writer: pd.ExcelWriter,
    *,
    engine: str,
    sheet_frames: Dict[str, pd.DataFrame],
) -> None:
    """افزودن ساختار Table به شیت‌های خروجی در صورت امکان."""

    if not sheet_frames:
        return

    table_names: set[str] = set()

    if engine == "xlsxwriter":
        for sheet_name, df in sheet_frames.items():
            worksheet = writer.sheets.get(sheet_name)
            if worksheet is None:
                continue
            rows, cols = df.shape
            if cols == 0:
                continue
            table_name = _build_table_name(sheet_name, table_names)
            worksheet.add_table(
                0,
                0,
                max(rows, 0),
                cols - 1,
                {
                    "name": table_name,
                    "columns": [{"header": header} for header in df.columns],
                    "style": None,
                },
            )
        return

    if engine != "openpyxl":
        return

    try:
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except Exception:  # pragma: no cover - وابستگی اختیاری
        return

    workbook = writer.book  # type: ignore[attr-defined]
    for sheet_name, df in sheet_frames.items():
        worksheet = workbook[sheet_name]
        if worksheet.max_column == 0:
            continue
        table_name = _build_table_name(sheet_name, table_names)
        end_row = max(worksheet.max_row, len(df) + 1)
        end_col = worksheet.max_column
        ref = f"A1:{get_column_letter(end_col)}{end_row}"
        table = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        worksheet.add_table(table)


def _apply_excel_formatting(
    writer: pd.ExcelWriter,
    *,
    engine: str,
    rtl: bool,
    font_name: str | None,
    sheet_frames: Dict[str, pd.DataFrame],
) -> None:
    """اعمال تنظیمات RTL و فونت پیش‌فرض برای خروجی Excel."""

    if not rtl and not font_name:
        return

    if engine == "xlsxwriter":
        workbook = writer.book  # type: ignore[attr-defined]
        fmt_options: dict[str, object] = {
            "align": "center",
            "valign": "vcenter",
        }
        if font_name:
            fmt_options["font_name"] = font_name
            fmt_options["font_size"] = 8
        fmt = workbook.add_format(fmt_options)
        for worksheet in writer.sheets.values():
            if rtl:
                worksheet.right_to_left()
            worksheet.set_column(0, 16384, None, fmt)
        return

    if engine != "openpyxl":
        return

    try:
        from openpyxl.styles import Alignment, Font
    except Exception:  # pragma: no cover - وابستگی اختیاری
        Font = None  # type: ignore[assignment]
        Alignment = None  # type: ignore[assignment]

    workbook = writer.book  # type: ignore[attr-defined]
    for sheet_name, df in sheet_frames.items():
        worksheet = workbook[sheet_name]
        if rtl:
            worksheet.sheet_view.rightToLeft = True
        if not font_name and Alignment is None:
            continue

        header_font = (
            Font(name=font_name, size=8)
            if font_name and Font is not None
            else None
        )
        cell_alignment = (
            Alignment(horizontal="center", vertical="center")
            if Alignment is not None
            else None
        )

        max_row = max(worksheet.max_row, len(df) + 1)
        max_col = worksheet.max_column
        for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                if header_font is not None:
                    cell.font = header_font
                if cell_alignment is not None:
                    cell.alignment = cell_alignment


@contextlib.contextmanager
def _temporary_file_path(*, suffix: str = "", directory: Path | str | None = None) -> Iterator[Path]:
    """مدیریت مسیر فایل موقتی با پاک‌سازی خودکار پس از اتمام کار."""

    fd, name = tempfile.mkstemp(suffix=suffix, dir=directory)
    os.close(fd)
    path = Path(name)
    try:
        yield path
    finally:
        path.unlink(missing_ok=True)


def _pick_engine() -> str | None:
    """انتخاب بهترین engine نصب‌شده برای نوشتن Excel."""

    forced = os.getenv("EXCEL_ENGINE")
    if forced in {"openpyxl", "xlsxwriter"}:
        return forced

    for engine in ("openpyxl", "xlsxwriter"):
        try:
            __import__(engine)
        except Exception:
            continue
        return engine
    return None


def write_xlsx_atomic(
    data_dict: Dict[str, pd.DataFrame],
    filepath: Path | str | PathLike[str],
    *,
    rtl: bool | None = None,
    font_name: str | None = None,
    header_mode: HeaderMode | None = None,
) -> None:
    """نوشتن امن و اتمیک Excel با مدیریت نام شیت و انتخاب engine."""

    target_path = Path(filepath)
    target_path.parent.mkdir(parents=True, exist_ok=True)

    policy = get_policy()
    if rtl is None:
        rtl = policy.excel.rtl
    if font_name is None:
        font_name = policy.excel.font_name
    if header_mode is None:
        header_mode = policy.excel.header_mode_write

    engine = _pick_engine()
    taken: set[str] = set()
    written_frames: Dict[str, pd.DataFrame] = {}

    processed_data: Dict[str, pd.DataFrame] = {}
    for sheet_name, df in data_dict.items():
        prepared = _prepare_dataframe_for_excel(df)
        if header_mode:
            prepared = canonicalize_headers(prepared, header_mode=header_mode)
        processed_data[sheet_name] = prepared

    if engine is None:
        warnings.warn(
            "No Excel engine available; falling back to CSV outputs.",
            RuntimeWarning,
            stacklevel=2,
        )
        for sheet_name, df in processed_data.items():
            safe_name = _safe_sheet_name(str(sheet_name), taken)
            csv_name = f"{target_path.stem}-{safe_name}.csv"
            df.to_csv(target_path.with_name(csv_name), index=False)
        return

    with _temporary_file_path(suffix=".xlsx", directory=target_path.parent) as tmp_path:
        with pd.ExcelWriter(tmp_path, engine=engine) as writer:
            for sheet_name, df in processed_data.items():
                safe_name = _safe_sheet_name(str(sheet_name), taken)
                df.to_excel(writer, sheet_name=safe_name, index=False)
                written_frames[safe_name] = df

            _apply_excel_tables(
                writer,
                engine=engine,
                sheet_frames=written_frames,
            )
            _apply_excel_formatting(
                writer,
                engine=engine,
                rtl=bool(rtl),
                font_name=font_name,
                sheet_frames=written_frames,
            )
        os.replace(tmp_path, target_path)


def read_excel_first_sheet(path: Path | str | PathLike[str]) -> pd.DataFrame:
    """خواندن شیت اول فایل Excel به‌صورت DataFrame."""

    source = Path(path)
    try:
        with pd.ExcelFile(source) as workbook:
            if not workbook.sheet_names:
                raise ValueError(f"هیچ شیتی در فایل {source} یافت نشد.")
            return workbook.parse(workbook.sheet_names[0], dtype={ALT_CODE_COLUMN: str})
    except FileNotFoundError as exc:
        raise FileNotFoundError(f"فایل یافت نشد: {source}") from exc
    except Exception as exc:  # pragma: no cover - سناریوهای پیش‌بینی‌نشده
        raise ValueError(f"خطا در خواندن فایل {source}: {exc}") from exc


def read_crosswalk_workbook(
    path: Path | str | PathLike[str],
) -> tuple[pd.DataFrame, pd.DataFrame | None]:
    """خواندن شیت‌های موردنیاز Crosswalk."""

    source = Path(path)
    sheet_groups = "پایه تحصیلی (گروه آزمایشی)"
    try:
        with pd.ExcelFile(source) as workbook:
            if sheet_groups not in workbook.sheet_names:
                raise ValueError(f"شیت «{sheet_groups}» در Crosswalk یافت نشد")
            dtype_map = {ALT_CODE_COLUMN: str}
            groups_df = workbook.parse(sheet_groups, dtype=dtype_map)
            synonyms_df = None
            if "Synonyms" in workbook.sheet_names:
                synonyms_df = workbook.parse("Synonyms", dtype=dtype_map)
            return groups_df, synonyms_df
    except FileNotFoundError as exc:
        raise FileNotFoundError(f"فایل Crosswalk یافت نشد: {source}") from exc
    except Exception as exc:  # pragma: no cover - سناریوهای پیش‌بینی‌نشده
        raise ValueError(f"خطا در باز کردن Crosswalk: {exc}") from exc

