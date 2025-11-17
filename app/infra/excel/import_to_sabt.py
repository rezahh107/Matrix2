"""زیرسیستم خروجی ImportToSabt مطابق Policy-First."""

from __future__ import annotations

import json
from collections import OrderedDict
from datetime import datetime
import hashlib
import re
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping, MutableMapping, NamedTuple, Sequence, Final

import pandas as pd

from app.core.common.columns import canonicalize_headers, ensure_series
from app.core.common.normalization import normalize_fa
from app.core.common.phone_rules import (
    normalize_landline_series,
    normalize_mobile,
)
from app.infra.excel._writer import ensure_text_columns
from app.core.pipeline import (
    REGISTRATION_STATUS_CANDIDATES,
    debug_registration_distribution,
    enrich_student_contacts,
)
from app.infra.excel.common import attach_contact_columns

GF_FIELD_TO_COL: Mapping[str, Sequence[str]] = {
    # اطلاعات هویتی دانش‌آموز
    "101": ("student_first_name", "student_name", "نام"),
    "102": ("student_last_name", "student_family_name", "نام خانوادگی"),
    "1": ("student_full_name", "نام و نام‌خانوادگی"),
    "3": ("student_father_name", "نام پدر"),
    "143": ("student_national_code", "student_national_id", "کد ملی"),
    "2": ("student_birth_date", "تاریخ تولد"),
    "92": ("student_gender", "جنسیت"),
    "93": ("student_educational_status", "وضعیت تحصیلی"),
    "98": ("student_foreign_national", "اتباع خارجی"),
    "4": ("student_art_school_status", "وضعیت هنرستان"),
    # اطلاعات تماس
    "20": ("student_mobile", "موبایل دانش‌آموز", "تلفن همراه"),
    "21": ("contact1_mobile", "موبایل رابط 1", "تلفن رابط 1"),
    "23": ("contact2_mobile", "موبایل رابط 2", "تلفن رابط 2"),
    "22": ("student_landline", "تلفن ثابت", "تلفن"),
    "24": ("contact1_relationship", "نسبت رابط 1"),
    "25": ("contact1_name", "نام رابط 1"),
    "26": ("contact2_relationship", "نسبت رابط 2"),
    "27": ("contact2_name", "نام رابط 2"),
    # وضعیت تحصیلی و مدرسه
    "73": ("student_exam_group", "گروه آزمایشی"),
    "31": ("student_major_code", "کد گروه/رشته"),
    "5": ("student_average", "معدل"),
    "30": ("student_school_code", "کد مدرسه 1"),
    "29": ("student_school_name", "نام مدرسه"),
    "94": ("student_center", "مرکز ثبت‌نام"),
    # ثبت‌نام و حکمت
    "75": ("student_registration_status", "وضعیت ثبت نام"),
    "76": ("student_hekmat_tracking_code", "کد رهگیری حکمت"),
    "97": ("student_hekmat_package_type", "نوع بسته حکمت"),
    "60": ("student_postal_code", "کدپستی", "کد پستی"),
    "61": ("student_postal_code_alias", "کدپستی جایگزین", "کد پستی جایگزین"),
    # سایر فیلدها
    "7": ("student_class_number", "شماره کلاس"),
    "8": ("student_seat_number", "شماره صندلی"),
    "96": ("student_konkur_quota", "سهمیه کنکور"),
    "39": ("suggested_mentor_id", "پشتیبان پیشنهادی"),
    "62": ("student_notes", "توضیحات"),
    # فیلدهای سیستمی
    "150": ("submission_source", "منبع ارسال"),
    "151": ("form_version", "sa_form_version"),
}

_DIGIT_TRANSLATION = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")
_HEADER_SPACE_PATTERN = re.compile(r"[\s\u200c\u200f]+")
_HEADER_PUNCT_PATTERN = re.compile(r"[\-/]+")

REGISTRATION_STATUS_LABELS: Final[dict[int, str]] = {
    0: "عادی",
    1: "بنیاد",
    3: "حکمت",
}

__all__ = [
    "GF_FIELD_TO_COL",
    "load_exporter_config",
    "prepare_allocation_export_frame",
    "build_sheet2_frame",
    "apply_alias_rule",
    "build_summary_frame",
    "build_errors_frame",
    "build_optional_sheet_frame",
    "ensure_template_workbook",
    "write_import_to_sabt_excel",
    "build_header_signature",
    "ImportToSabtExportError",
    "map_registration_status_column",
]


class ExporterConfigError(ValueError):
    """خطای اعتبارسنجی تنظیمات Exporter."""


class ImportToSabtExportError(ValueError):
    """خطای اعتبارسنجی داده برای خروجی ImportToSabt."""


class HeaderMismatchLog(NamedTuple):
    """ثبت ساختاریافتهٔ مغایرت هدر Template با ستون‌های انتظاری."""

    sheet_name: str
    template_headers: tuple[str, ...]
    expected_headers: tuple[str, ...]
    expected_signature: str


def _describe_frame_source(frame: pd.DataFrame, *, default_label: str) -> str:
    """تولید توضیح خوانا از منبع دیتافریم برای گزارش خطا."""

    attrs = getattr(frame, "attrs", {}) or {}
    sheet = attrs.get("sheet_name") or attrs.get("sheet_label") or attrs.get("source_sheet")
    source_label = attrs.get("source_label") or attrs.get("source") or attrs.get("origin")
    path = attrs.get("source_path") or attrs.get("path") or attrs.get("filepath")

    label_parts: list[str] = []
    if sheet:
        label_parts.append(f"sheet '{sheet}'")
    if source_label and source_label not in label_parts:
        label_parts.append(str(source_label))
    if path:
        label_parts.append(str(path))

    if label_parts:
        if len(label_parts) == 1:
            return label_parts[0]
        if len(label_parts) == 2 and label_parts[0].startswith("sheet"):
            return f"{label_parts[0]} from {label_parts[1]}"
        return " | ".join(label_parts)

    return default_label


def validate_export_identifiers(
    frame: pd.DataFrame,
    *,
    required_column: str,
    entity_name: str,
    source_label: str,
    alternate_column: str | None = None,
) -> None:
    """بررسی وجود ستون شناسه پیش از نرمال‌سازی و تولید پیام راهنما."""

    candidates = [required_column]
    if alternate_column is not None:
        candidates.append(alternate_column)

    for column in candidates:
        if column in frame.columns:
            return

    column_hint = " یا ".join(f"'{col}'" for col in candidates)
    raise ImportToSabtExportError(
        "ImportToSabt export failed: missing required "
        f"{entity_name} identifier column ({column_hint}) in {source_label}. "
        "Please add the column to the source sheet/file and rerun the exporter."
    )


def load_exporter_config(path: str | Path) -> dict:
    """بارگذاری config با حفظ ترتیب ستون‌ها و اعتبارسنجی حداقلی."""

    cfg_path = Path(path)
    if not cfg_path.exists():
        raise FileNotFoundError(f"Exporter config not found: {cfg_path}")

    with cfg_path.open("r", encoding="utf-8") as handle:
        cfg = json.load(handle, object_pairs_hook=OrderedDict)

    sheets = cfg.get("sheets")
    if not isinstance(sheets, MutableMapping):
        raise ExporterConfigError("'sheets' must be a dict in exporter config")

    sheet2 = sheets.get("Sheet2")
    if not isinstance(sheet2, MutableMapping):
        raise ExporterConfigError("Exporter config must define 'Sheet2'")

    columns = sheet2.get("columns")
    if not isinstance(columns, MutableMapping) or not columns:
        raise ExporterConfigError("'Sheet2.columns' must be a non-empty dict")

    return cfg


def prepare_allocation_export_frame(
    allocations_df: pd.DataFrame,
    students_df: pd.DataFrame,
    mentors_df: pd.DataFrame,
    *,
    student_ids: pd.Series | None = None,
) -> pd.DataFrame:
    """ادغام ستون‌های دانش‌آموز و پشتیبان روی دیتافریم تخصیص."""

    alloc = canonicalize_headers(allocations_df, header_mode="en").copy()
    status_debug: list[dict[str, Any]] = []

    def _capture_status(label: str, frame: pd.DataFrame) -> None:
        status_debug.append(
            {"label": label, **debug_registration_distribution(frame, REGISTRATION_STATUS_CANDIDATES)}
        )

    _capture_status("students_raw", students_df)

    students_contacts = enrich_student_contacts(students_df)
    _capture_status("students_contacts", students_contacts)

    students = canonicalize_headers(students_df, header_mode="en").copy()
    _capture_status("students_canonical", students)

    students = attach_contact_columns(students, students_contacts)
    _capture_status("students_with_contacts", students)
    mentors = canonicalize_headers(mentors_df, header_mode="en").copy()

    dedupe_logs: list[dict[str, list[str]]] = []
    dedupe_cache: dict[str, pd.DataFrame] = {}

    def _register_dedupe_log(context: str, frame: pd.DataFrame) -> None:
        removed = _get_deduplicated_columns(frame, context=context)
        if removed:
            dedupe_logs.append({"context": context, "columns": removed})

    def _dedupe_cached(frame: pd.DataFrame, context: str) -> pd.DataFrame:
        cached = dedupe_cache.get(context)
        if cached is not None:
            return cached
        if not frame.columns.duplicated().any():
            dedupe_cache[context] = frame
            return frame
        deduped = _deduplicate_columns(frame, context=context)
        _register_dedupe_log(context, deduped)
        dedupe_cache[context] = deduped
        return deduped

    alloc = _dedupe_cached(alloc, "allocations")
    students = _dedupe_cached(students, "students_source")
    mentors = _dedupe_cached(mentors, "mentors_source")

    student_source = _describe_frame_source(students_df, default_label="students_df")
    mentor_source = _describe_frame_source(mentors_df, default_label="mentors_df")

    if student_ids is not None:
        aligned_ids = student_ids.reindex(students.index)
        students.loc[:, "student_id"] = aligned_ids.astype("string")

    validate_export_identifiers(
        students,
        required_column="student_id",
        entity_name="student",
        source_label=student_source,
    )
    validate_export_identifiers(
        mentors,
        required_column="mentor_id",
        entity_name="mentor",
        source_label=mentor_source,
        alternate_column="alias",
    )

    if "mentor_id" not in mentors.columns:
        alias_series = mentors.get("alias")
        if alias_series is not None:
            mentors.loc[:, "mentor_id"] = alias_series

    def _prefix(frame: pd.DataFrame, prefix: str, preserve: Iterable[str]) -> pd.DataFrame:
        renamed: dict[str, str] = {}
        for column in frame.columns:
            name = str(column)
            if name in preserve:
                renamed[name] = name
            else:
                renamed[name] = f"{prefix}{name}"
        return frame.rename(columns=renamed)

    students_prefixed = _prefix(students, "student_", {"student_id"})
    mentors_prefixed = _prefix(mentors, "mentor_", {"mentor_id"})
    mentors_prefixed = _coalesce_duplicate_identifier_rows(
        mentors_prefixed,
        "mentor_id",
        entity_name="mentor",
    )

    students_prefixed = _dedupe_cached(students_prefixed, "students")
    mentors_prefixed = _dedupe_cached(mentors_prefixed, "mentors")

    _ensure_unique_identifier(students_prefixed, "student_id", entity_name="student")
    _ensure_unique_identifier(mentors_prefixed, "mentor_id", entity_name="mentor")

    merged = _safe_merge(
        alloc,
        students_prefixed,
        how="left",
        on="student_id",
        sort=False,
        validate="many_to_one",
        context="student",
        left_label="allocations dataframe",
        right_label="student dataframe",
    )
    merged = _safe_merge(
        merged,
        mentors_prefixed,
        how="left",
        on="mentor_id",
        sort=False,
        validate="many_to_one",
        context="mentor",
        left_label="allocations dataframe",
        right_label="mentor dataframe",
    )

    if len(merged.index) != len(alloc.index):
        raise ImportToSabtExportError(
            "ImportToSabt export failed: expected "
            f"{len(alloc)} allocation rows after joining student/mentor details, but got "
            f"{len(merged)}. This usually means duplicate student_id or mentor_id records "
            "exist in the exporter inputs (مثلاً فایل استخر یا دانش‌آموز اشتباه). "
            "Please fix the input data and retry."
        )

    merged.index = alloc.index
    if dedupe_logs:
        merged.attrs["dedupe_logs"] = dedupe_logs
    _capture_status("merged_before_enrich", merged)
    merged = enrich_student_contacts(merged)
    _capture_status("merged_after_enrich", merged)
    merged.attrs["registration_status_debug"] = status_debug
    return merged


def _ensure_unique_identifier(frame: pd.DataFrame, column: str, *, entity_name: str) -> None:
    """اطمینان از یکتایی شناسه در دیتافریم ورودی برای جلوگیری از join چند به چند."""

    if column not in frame.columns:
        raise ImportToSabtExportError(
            f"ImportToSabt export failed: '{column}' column not found for {entity_name} data."
        )

    series = frame[column]
    non_na = series.dropna()
    duplicated_ids = non_na[non_na.duplicated(keep=False)]
    if not duplicated_ids.empty:
        sample = ", ".join(duplicated_ids.astype(str).unique()[:5])
        raise ImportToSabtExportError(
            "ImportToSabt export failed: detected duplicate "
            f"{entity_name}_id values ({sample}). Each {entity_name}_id must be unique "
            "in ImportToSabt exporter inputs. لطفاً داده ورودی را اصلاح کنید."
        )


def _coalesce_duplicate_identifier_rows(
    frame: pd.DataFrame,
    column: str,
    *,
    entity_name: str,
) -> pd.DataFrame:
    """ترکیب ردیف‌های تکراری بر اساس شناسه با حفظ ترتیب و پرکردن مقادیر خالی."""

    if column not in frame.columns:
        raise ImportToSabtExportError(
            f"ImportToSabt export failed: '{column}' column not found for {entity_name} data."
        )

    normalized = ensure_series(frame[column]).astype("string").fillna("").str.strip()
    duplicate_mask = normalized.duplicated(keep=False) & normalized.ne("")
    if not bool(duplicate_mask.any()):
        return frame

    merged_lookup: dict[str, dict[str, Any]] = {}
    records: list[dict[str, Any]] = []
    normalized_values = normalized.to_numpy()
    for identifier, (_, row) in zip(normalized_values, frame.iterrows()):
        if pd.isna(identifier) or str(identifier).strip() == "":
            records.append(row.to_dict())
            continue
        key = str(identifier).strip()
        existing = merged_lookup.get(key)
        if existing is None:
            data = row.to_dict()
            merged_lookup[key] = data
            records.append(data)
            continue
        for field, value in row.items():
            if _is_missing_value(existing.get(field)) and not _is_missing_value(value):
                existing[field] = value

    return pd.DataFrame(records, columns=frame.columns)


_DEDUPLICATED_COLUMNS_ATTR = "_deduplicated_columns"


def _deduplicate_columns(frame: pd.DataFrame, *, context: str) -> pd.DataFrame:
    """حذف ستون‌های تکراری با حفظ ترتیب و تشخیص داده‌های ناسازگار."""

    columns = frame.columns
    if columns.is_unique:
        return frame

    duplicated_names = columns[columns.duplicated()].unique().tolist()
    conflicts: list[str] = []
    for name in duplicated_names:
        subset = frame.loc[:, frame.columns == name]
        base = ensure_series(subset.iloc[:, 0])
        for idx in range(1, subset.shape[1]):
            candidate = ensure_series(subset.iloc[:, idx])
            if not _series_semantically_equal(base, candidate):
                conflicts.append(str(name))
                break
    if conflicts:
        sample = ", ".join(conflicts[:5])
        raise ImportToSabtExportError(
            "ImportToSabt export failed: duplicate columns with conflicting data "
            f"detected for {context} dataframe ({sample})."
        )

    mask = ~columns.duplicated(keep="first")
    result = frame.loc[:, mask].copy()
    _store_deduplicated_columns(result, context=context, removed=duplicated_names)
    return result


def _store_deduplicated_columns(
    frame: pd.DataFrame, *, context: str, removed: Sequence[str]
) -> None:
    """ذخیرهٔ ستون‌های حذف‌شده در attrs برای گزارش‌گیری."""

    if not removed:
        return
    attrs = getattr(frame, "attrs", None)
    if attrs is None:
        return
    registry = attrs.get(_DEDUPLICATED_COLUMNS_ATTR)
    if not isinstance(registry, dict):
        registry = {}
        attrs[_DEDUPLICATED_COLUMNS_ATTR] = registry
    registry[context] = [str(name) for name in removed]


def _get_deduplicated_columns(frame: pd.DataFrame, *, context: str) -> list[str]:
    """دریافت ستون‌های حذف‌شده برای context مشخص."""

    attrs = getattr(frame, "attrs", None)
    if not attrs:
        return []
    registry = attrs.get(_DEDUPLICATED_COLUMNS_ATTR)
    if not isinstance(registry, dict):
        return []
    removed = registry.get(context, [])
    if isinstance(removed, list):
        return [str(name) for name in removed]
    return []


def _series_semantically_equal(left: pd.Series, right: pd.Series) -> bool:
    """مقایسهٔ دو ستون با نادیده گرفتن اختلاف تایپ یا صفرهای پیشرو."""

    left_series = ensure_series(left)
    right_series = ensure_series(right)
    if len(left_series) != len(right_series):
        return False

    left_numeric = pd.to_numeric(left_series, errors="coerce")
    right_numeric = pd.to_numeric(right_series, errors="coerce")
    if left_numeric.notna().any() or right_numeric.notna().any():
        if left_numeric.equals(right_numeric):
            return True

    left_normalized = left_series.astype("string").fillna("").str.strip()
    right_normalized = right_series.astype("string").fillna("").str.strip()
    return left_normalized.equals(right_normalized)


def _is_missing_value(value: Any) -> bool:
    """تشخیص مقدار خالی با درنظر گرفتن رشتهٔ تهی و NaN."""

    if isinstance(value, str):
        return value.strip() == ""
    return pd.isna(value)


_MERGE_DUPLICATE_SAMPLE_LIMIT = 5


def _safe_merge(
    left: pd.DataFrame,
    right: pd.DataFrame,
    *,
    context: str,
    left_label: str | None = None,
    right_label: str | None = None,
    sample_limit: int = _MERGE_DUPLICATE_SAMPLE_LIMIT,
    **kwargs: Any,
) -> pd.DataFrame:
    """اجرای merge امن با پیام خطای قابل‌اقدام.

    مثال:
        >>> left = pd.DataFrame({"student_id": ["A", "B"]})
        >>> right = pd.DataFrame({"student_id": ["A", "A"]})
        >>> _safe_merge(left, right, context="student", on="student_id", validate="one_to_one")
        Traceback (most recent call last):
            ... ImportToSabtExportError: ImportToSabt export failed ... duplicate keys ...
    """

    left_name = left_label or _describe_frame_source(left, default_label="left dataframe")
    right_name = right_label or _describe_frame_source(right, default_label="right dataframe")

    try:
        return left.merge(right, **kwargs)
    except ValueError as exc:  # pragma: no cover - مسیر خطا تست دارد
        duplicate_hint = _format_merge_duplicate_hint(
            str(exc),
            left,
            right,
            left_name,
            right_name,
            sample_limit=sample_limit,
            merge_kwargs=kwargs,
        )
        raise ImportToSabtExportError(
            "ImportToSabt export failed while joining "
            f"{context} details: {exc}{duplicate_hint}"
        ) from exc


def _format_merge_duplicate_hint(
    message: str,
    left: pd.DataFrame,
    right: pd.DataFrame,
    left_label: str,
    right_label: str,
    *,
    sample_limit: int,
    merge_kwargs: Mapping[str, Any],
) -> str:
    """تولید پیام کمکی برای خطای merge با نمایش شناسه‌های تکراری.

    مثال:
        >>> left = pd.DataFrame({"id": ["A", "A"]})
        >>> right = pd.DataFrame({"id": ["A", "A"]})
        >>> _format_merge_duplicate_hint(
        ...     "Merge keys are not unique in right dataset", left, right, "left", "right",
        ...     sample_limit=2,
        ...     merge_kwargs={"on": "id"},
        ... )
        " Sample duplicate keys detected (right frame 'right' duplicate keys: A)."
    """

    normalized_message = message.lower()
    if "merge keys are not unique" not in normalized_message:
        return ""

    sides: set[str] = set()
    if "left dataset" in normalized_message:
        sides.add("left")
    if "right dataset" in normalized_message:
        sides.add("right")
    if not sides and "either left or right" in normalized_message:
        sides = {"left", "right"}

    join_config = _resolve_merge_join_config(merge_kwargs)
    left_keys = _build_merge_key_series(
        left,
        columns=join_config.left_on,
        use_index=join_config.left_index,
    )
    right_keys = _build_merge_key_series(
        right,
        columns=join_config.right_on,
        use_index=join_config.right_index,
    )

    notes: list[str] = []
    if "left" in sides:
        left_samples, left_duplicate_set = _collect_duplicate_key_samples(left_keys, sample_limit)
        if left_samples:
            matching_right = _matching_duplicate_keys(right_keys, left_duplicate_set, sample_limit)
            note = f"left frame '{left_label}' duplicate keys: {', '.join(left_samples)}"
            if matching_right:
                note += f" | matching right frame '{right_label}': {', '.join(matching_right)}"
            notes.append(note)
    if "right" in sides:
        right_samples, right_duplicate_set = _collect_duplicate_key_samples(right_keys, sample_limit)
        if right_samples:
            matching_left = _matching_duplicate_keys(left_keys, right_duplicate_set, sample_limit)
            note = f"right frame '{right_label}' duplicate keys: {', '.join(right_samples)}"
            if matching_left:
                note += f" | matching left frame '{left_label}': {', '.join(matching_left)}"
            notes.append(note)

    if not notes:
        return ""
    return f" Sample duplicate keys detected ({'; '.join(notes)})."


class _MergeJoinConfig(NamedTuple):
    left_on: list[str] | None
    right_on: list[str] | None
    left_index: bool
    right_index: bool


def _resolve_merge_join_config(kwargs: Mapping[str, Any]) -> _MergeJoinConfig:
    """استخراج تنظیمات join برای تحلیل خطا.

    مثال:
        >>> _resolve_merge_join_config({"on": "student_id"})
        _MergeJoinConfig(left_on=['student_id'], right_on=['student_id'], left_index=False, right_index=False)
    """

    on_value = kwargs.get("on")
    left_index = bool(kwargs.get("left_index"))
    right_index = bool(kwargs.get("right_index"))

    if on_value is not None:
        columns = [on_value] if isinstance(on_value, str) else list(on_value)
        return _MergeJoinConfig(columns, columns, left_index, right_index)

    left_on = kwargs.get("left_on")
    right_on = kwargs.get("right_on")
    left_columns = [left_on] if isinstance(left_on, str) else (list(left_on) if left_on is not None else None)
    right_columns = [right_on] if isinstance(right_on, str) else (list(right_on) if right_on is not None else None)
    return _MergeJoinConfig(left_columns, right_columns, left_index, right_index)


def _build_merge_key_series(
    frame: pd.DataFrame,
    *,
    columns: Sequence[str] | None,
    use_index: bool,
) -> pd.Series:
    """ساخت سری کلیدهای join با رشته‌سازی پایدار.

    مثال:
        >>> df = pd.DataFrame({"id": ["A", "B"], "code": [1, 2]})
        >>> _build_merge_key_series(df, columns=["id"], use_index=False).tolist()
        ['A', 'B']
    """

    parts: list[pd.Series] = []
    if use_index:
        index_series = pd.Series(frame.index, index=frame.index, name=frame.index.name or "index")
        parts.append(index_series)

    if columns:
        existing_columns = [col for col in columns if col in frame.columns]
        for column in existing_columns:
            parts.append(ensure_series(frame[column]))

    if not parts:
        return pd.Series([], dtype="string")

    normalized = [series.astype("string").fillna("") for series in parts]
    if len(normalized) == 1:
        return normalized[0]

    data = pd.concat(normalized, axis=1)
    return data.apply(lambda row: " | ".join(row.astype(str)), axis=1)


def _collect_duplicate_key_samples(series: pd.Series, sample_limit: int) -> tuple[list[str], set[str]]:
    """استخراج نمونه شناسه‌های تکراری از یک سری join.

    مثال:
        >>> series = pd.Series(["A", "A", "B"])
        >>> _collect_duplicate_key_samples(series, 2)
        (['A'], {'A'})
    """

    if series.empty:
        return [], set()

    duplicated_mask = series.duplicated(keep=False)
    if not bool(duplicated_mask.any()):
        return [], set()

    duplicates = series[duplicated_mask]
    seen: set[str] = set()
    samples: list[str] = []
    for value in duplicates:
        key = _stringify_merge_key(value)
        if key in seen:
            continue
        seen.add(key)
        samples.append(key)
        if len(samples) >= sample_limit:
            break
    return samples, {_stringify_merge_key(value) for value in duplicates}


def _matching_duplicate_keys(
    series: pd.Series,
    duplicates: set[str],
    sample_limit: int,
) -> list[str]:
    """برگرداندن نمونه شناسه‌های مشترک بین سری و مجموعهٔ تکراری.

    مثال:
        >>> series = pd.Series(["A", "B", "C"])
        >>> _matching_duplicate_keys(series, {"A", "C"}, 1)
        ['A']
    """

    if not duplicates or series.empty:
        return []

    matches: list[str] = []
    seen: set[str] = set()
    for value in series:
        key = _stringify_merge_key(value)
        if key in duplicates and key not in seen:
            seen.add(key)
            matches.append(key)
            if len(matches) >= sample_limit:
                break
    return matches


def _stringify_merge_key(value: Any) -> str:
    """تبدیل مقدار کلید join به متن پایدار برای گزارش خطا.

    مثال:
        >>> _stringify_merge_key('')
        '<EMPTY>'
    """

    if value is None:
        return "<EMPTY>"
    text = str(value).strip()
    return text if text else "<EMPTY>"


def _normalize_digits(value: Any, length: int) -> str:
    text = str(value or "").translate(_DIGIT_TRANSLATION)
    digits = "".join(ch for ch in text if ch.isdigit())
    if not digits:
        return ""
    digits = digits[-length:].zfill(length)
    return digits[:length]


def _apply_normalizers(series: pd.Series, normalizer: Any) -> pd.Series:
    """اعمال نرمال‌سازهای تعریف‌شده روی یک سری داده."""

    if normalizer is None:
        return series

    items = normalizer if isinstance(normalizer, (list, tuple)) else [normalizer]
    result = series.astype("object").copy()

    for name in items:
        if name == "digits_10":
            result = result.map(lambda v: _normalize_digits(v, 10))
        elif name == "digits_16":
            result = result.map(lambda v: _normalize_digits(v, 16))
        elif name == "mobile_ir":
            result = result.map(_normalize_mobile_ir)
        else:
            raise ValueError(f"Unknown normalizer: {name!r}")

    return result


def _normalize_mobile_ir(value: Any) -> str:
    """نرمال‌سازی موبایل از طریق :func:`app.core.common.phone_rules.normalize_mobile`."""

    normalized = normalize_mobile(value)
    return normalized or ""


def _resolve_map(map_spec: Any, cfg: Mapping[str, Any]) -> Mapping[Any, Any] | None:
    if map_spec is None:
        return None
    if isinstance(map_spec, str):
        maps = cfg.get("maps")
        if isinstance(maps, Mapping) and map_spec in maps:
            return maps[map_spec]
    if isinstance(map_spec, Mapping):
        return map_spec
    return None


def _coerce_type(series: pd.Series, type_name: str | None, precision: int | None = None) -> pd.Series:
    if type_name == "number":
        numeric = pd.to_numeric(series, errors="coerce")
        if precision is not None:
            numeric = numeric.round(precision)
        return numeric
    converted = series.astype("string")
    return converted.fillna("")


def _normalize_date_format(fmt: str | None) -> str:
    if not fmt:
        return "%Y-%m-%d"
    normalized = str(fmt)
    replacements = {
        "yyyy": "%Y",
        "yy": "%y",
        "mm": "%m",
        "dd": "%d",
        "HH": "%H",
        "MM": "%M",
        "ss": "%S",
    }
    for token, replacement in replacements.items():
        normalized = normalized.replace(token, replacement)
    return normalized


def _series_from_source(
    df: pd.DataFrame,
    source_cfg: Mapping[str, Any],
    exporter_cfg: Mapping[str, Any],
    *,
    today: datetime,
) -> pd.Series:
    source_type = source_cfg.get("source", "df")
    sheet_cfg = exporter_cfg.get("sheets", {}).get("Sheet2", {})
    columns_cfg = sheet_cfg.get("columns", {}) if isinstance(sheet_cfg, Mapping) else {}

    def _empty_series() -> pd.Series:
        return pd.Series(["" for _ in range(len(df))], index=df.index, dtype="string")

    if source_type == "empty":
        return _empty_series()

    if source_type == "literal":
        value = source_cfg.get("value", "")
        return pd.Series([value for _ in range(len(df))], index=df.index)

    if source_type == "system":
        field = source_cfg.get("field")
        fmt = _normalize_date_format(source_cfg.get("format"))
        if field == "date_created":
            for column_name in ("student_created_at", "created_at"):
                if column_name in df.columns:
                    series = pd.to_datetime(df[column_name], errors="coerce")
                    formatted = series.dt.strftime(fmt)
                    return formatted.fillna("")
            return pd.Series([today.strftime(fmt) for _ in range(len(df))], index=df.index)
        if field == "today":
            return pd.Series([today.strftime(fmt) for _ in range(len(df))], index=df.index)
        return pd.Series([source_cfg.get("value", "") for _ in range(len(df))], index=df.index)

    if source_type == "gf":
        field = source_cfg.get("field")
        candidates: list[str] = []
        if isinstance(field, str):
            mapped = GF_FIELD_TO_COL.get(field)
            if mapped:
                candidates.extend(mapped)
            prefixed = f"student_{field}" if not field.startswith("student_") else field
            candidates.extend([prefixed, f"student_gf_{field}", field])
        expanded_candidates: list[str] = []
        for column_name in candidates:
            if not column_name:
                continue
            expanded_candidates.append(column_name)
            expanded_candidates.append(f"student_{column_name}")
        candidates = expanded_candidates
        for column_name in dict.fromkeys(filter(None, candidates)):
            if column_name in df.columns:
                series = df[column_name]
                fallback_column = source_cfg.get("on_value_9000_use")
                if fallback_column:
                    if fallback_column in df.columns:
                        fallback_series = df[fallback_column]
                    else:
                        fallback_series = None
                        fallback_spec = columns_cfg.get(fallback_column)
                        if isinstance(fallback_spec, Mapping) and fallback_spec is not source_cfg:
                            fallback_series = _series_from_source(
                                df,
                                fallback_spec,
                                exporter_cfg,
                                today=today,
                            )
                    if fallback_series is not None:
                        mask = series.astype("string") == "9000"
                        series = series.copy()
                        fallback_values = ensure_series(fallback_series).reindex(df.index).fillna("")
                        series.loc[mask] = fallback_values[mask]
                return series
        return _empty_series()

    if source_type == "derived":
        from_column = source_cfg.get("from")
        crosswalk_name = source_cfg.get("crosswalk")
        base_series: pd.Series | None = None
        if isinstance(from_column, str):
            if from_column in df.columns:
                base_series = df[from_column]
            elif isinstance(columns_cfg, Mapping):
                from_spec = columns_cfg.get(from_column)
                if isinstance(from_spec, Mapping) and from_spec is not source_cfg:
                    base_series = _series_from_source(
                        df,
                        from_spec,
                        exporter_cfg,
                        today=today,
                    )
        if base_series is None:
            return _empty_series()
        if crosswalk_name:
            lookups = exporter_cfg.get("lookups", {})
            crosswalk = lookups.get(crosswalk_name) if isinstance(lookups, Mapping) else None
            if isinstance(crosswalk, Mapping):
                mapped = ensure_series(base_series).map(crosswalk)
                return mapped.fillna("")
        return ensure_series(base_series)

    if source_type == "lookup":
        lookup_name = source_cfg.get("lookup")
        column = source_cfg.get("field")
        lookups = exporter_cfg.get("lookups", {})
        table = lookups.get(lookup_name) if isinstance(lookups, Mapping) else None
        if isinstance(table, Mapping) and column in df.columns:
            series = df[column].map(table)
            return series.fillna("")
        return _empty_series()

    column = source_cfg.get("field")
    if column in df.columns:
        return df[column]
    return _empty_series()


def build_sheet2_frame(
    df_alloc: pd.DataFrame,
    exporter_cfg: Mapping[str, Any],
    today: datetime | None = None,
) -> pd.DataFrame:
    """ساخت دیتافریم Sheet2 بر اساس تنظیمات JSON."""

    if today is None:
        today = datetime.today()
    debug_log: list[dict[str, Any]] = []

    existing_debug = df_alloc.attrs.get("registration_status_debug")
    if isinstance(existing_debug, list):
        debug_log.extend(existing_debug)

    debug_log.append(
        {
            "label": "sheet2_input",
            **debug_registration_distribution(df_alloc, REGISTRATION_STATUS_CANDIDATES),
        }
    )

    df_alloc = enrich_student_contacts(df_alloc)
    debug_log.append(
        {
            "label": "sheet2_after_enrich",
            **debug_registration_distribution(df_alloc, REGISTRATION_STATUS_CANDIDATES),
        }
    )
    if "student_registration_status" not in df_alloc.columns:
        df_alloc["student_registration_status"] = pd.Series(
            [pd.NA] * len(df_alloc), index=df_alloc.index, dtype="Int64"
        )
    sheet_cfg = exporter_cfg["sheets"]["Sheet2"]
    columns_cfg = sheet_cfg["columns"]
    if isinstance(columns_cfg, OrderedDict):
        ordered_columns = list(columns_cfg.keys())
    elif isinstance(columns_cfg, Mapping):
        ordered_columns = list(columns_cfg.keys())
    else:
        raise TypeError("'columns' config must be a mapping")

    sheet = pd.DataFrame(index=df_alloc.index)
    for column_name in ordered_columns:
        spec = columns_cfg[column_name]
        series = _series_from_source(df_alloc, spec, exporter_cfg, today=today)
        if not isinstance(series, pd.Series):
            series = pd.Series(series, index=df_alloc.index)
        series = series.reindex(df_alloc.index)
        series = _apply_normalizers(series, spec.get("normalize"))
        map_dict = _resolve_map(spec.get("map"), exporter_cfg)

        if column_name == "وضعیت ثبت نام":
            debug_log.append(
                {
                    "label": "sheet2_status_raw_series",
                    **debug_registration_distribution(
                        pd.DataFrame({"candidate": series}), ("candidate",)
                    ),
                }
            )
            normalized_status = _normalize_registration_status(series)
            debug_log.append(
                {
                    "label": "sheet2_status_normalized",
                    **debug_registration_distribution(
                        pd.DataFrame({"candidate": normalized_status}), ("candidate",)
                    ),
                }
            )
            series = map_registration_status_column(normalized_status)
            map_dict = None

        if map_dict:
            mapped = series.astype("string").map(map_dict)
            series = mapped.fillna(series)

        series = _coerce_type(series, spec.get("type"), spec.get("precision"))

        when_clause = spec.get("when")
        if when_clause:
            parts = str(when_clause).split("=")
            if len(parts) == 2:
                condition_col = parts[0].strip()
                condition_val = parts[1].strip()
                if condition_col in sheet.columns:
                    targets = {condition_val}
                    condition_spec = columns_cfg.get(condition_col)
                    if isinstance(condition_spec, Mapping):
                        cond_map = _resolve_map(condition_spec.get("map"), exporter_cfg)
                        if cond_map and condition_val in cond_map:
                            targets.add(cond_map[condition_val])
                    mask = sheet[condition_col].astype("string").isin(targets)
                    inactive = ~mask.fillna(False)
                    if inactive.any():
                        series = series.copy()
                        series.loc[inactive] = ""

        sheet[column_name] = series

    sheet = sheet.loc[:, ordered_columns]

    landline_column = sheet_cfg.get("landline_column") or "تلفن ثابت"
    if landline_column in sheet.columns:
        landline_source = df_alloc.get("student_landline")
        if landline_source is not None:
            aligned_landline = ensure_series(landline_source).reindex(df_alloc.index)
            sheet[landline_column] = aligned_landline.astype("string").fillna("")
        sheet[landline_column] = normalize_landline_series(
            sheet[landline_column], allow_special_zero=True
        ).fillna("")

    sheet = sheet.astype({column: "string" for column in sheet.columns})
    sheet.attrs["exporter_config"] = exporter_cfg
    sheet.attrs["registration_status_debug"] = debug_log
    return sheet


def _normalize_registration_status(series: pd.Series) -> pd.Series:
    """بازگرداندن ستون وضعیت ثبت‌نام با حفظ مقادیر ۰/۱/۳.

    این تابع فقط جایگزین امن برای مقادیر خالی است و مقدار موجود را تغییر نمی‌دهد.
    ارقام فارسی/عربی را به رقم لاتین تبدیل می‌کند تا مقادیر ۳ به درستی شناسایی شوند.

    مثال::
        >>> s = pd.Series([0, 3, None])
        >>> _normalize_registration_status(s).tolist()
        [0, 3, 0]
    """

    normalized = ensure_series(series)
    as_string = normalized.astype("string")
    translated = as_string.str.translate(_DIGIT_TRANSLATION).str.strip()
    numeric = pd.to_numeric(translated, errors="coerce")
    blank_mask = numeric.isna()
    filled = numeric.mask(blank_mask, 0)
    return filled.astype("Int64")


def map_registration_status_column(series: pd.Series) -> pd.Series:
    """نگاشت امن وضعیت ثبت‌نام از کد عددی به متن فارسی.

    ورودی باید شامل کدهای ۰/۱/۳ (Int64 یا قابل تبدیل) باشد. مقادیر خالی یا
    نامعتبر به صورت پیش‌فرض «عادی» می‌شوند تا خروجی شفاف و پایدار باشد.

    مثال::
        >>> s = pd.Series([0, 3, None], dtype="Int64")
        >>> map_registration_status_column(s).tolist()
        ['عادی', 'حکمت', 'عادی']
    """

    numeric = _normalize_registration_status(series)
    mapped = numeric.map(REGISTRATION_STATUS_LABELS)
    return mapped.fillna("عادی")


def apply_alias_rule(sheet2: pd.DataFrame, df_alloc: pd.DataFrame) -> pd.DataFrame:
    """اعمال قانون alias برای ستون‌های کد پستی و کد پستی جایگزین."""

    cfg = sheet2.attrs.get("exporter_config") or {}
    alias_cfg = cfg.get("alias_rule") if isinstance(cfg, Mapping) else None

    def _first_series(candidates: Iterable[str | None]) -> pd.Series | None:
        for name in candidates:
            if name and name in df_alloc.columns:
                return df_alloc[name]
        return None

    if isinstance(alias_cfg, Mapping) and alias_cfg:
        sheet_postal = alias_cfg.get("sheet_postal_column")
        df_postal = alias_cfg.get("df_postal_column")
        if not sheet_postal or not df_postal or sheet_postal not in sheet2.columns:
            return sheet2
        source_postal = df_alloc.get(df_postal)
        if source_postal is None:
            return sheet2
        normalized_postal = source_postal.astype("string").fillna("")
        normalized_postal = normalized_postal.map(lambda v: _normalize_digits(v, 10))
        sheet2.loc[:, sheet_postal] = normalized_postal

        sheet_alias = alias_cfg.get("sheet_alias_column")
        df_alias = alias_cfg.get("df_alias_column")
        alias_series = df_alloc.get(df_alias) if df_alias else None
        if alias_series is not None:
            alias_series = alias_series.astype("string").fillna("")
            alias_series = alias_series.map(lambda v: _normalize_digits(v, 10))
            if sheet_alias and sheet_alias in sheet2.columns:
                sheet2.loc[:, sheet_alias] = alias_series.reindex(df_alloc.index)
            mask_alias = alias_series.str.strip() != ""
            if mask_alias.any():
                sheet2.loc[mask_alias, sheet_postal] = alias_series[mask_alias]
        return sheet2

    sheet_postal = "کد پستی"
    sheet_alias = "کد پستی جایگزین"

    if sheet_postal not in sheet2.columns:
        return sheet2

    mentor_alias = _first_series(
        ["mentor_alias_code", "mentor_alias_postal_code", "mentor_postal_code"]
    )
    if mentor_alias is not None:
        alias_series = mentor_alias.astype("string").fillna("")
        alias_series = alias_series.map(lambda v: _normalize_digits(v, 10))
        mask_alias = alias_series.str.strip() != ""
        if mask_alias.any():
            sheet2.loc[mask_alias, sheet_postal] = alias_series[mask_alias]

    mentor_id_series = _first_series(["mentor_id", "mentor_mentor_id"])
    school_limit_series = _first_series(["mentor_is_school_limited", "is_school_limited"])

    if (
        sheet_alias in sheet2.columns
        and mentor_id_series is not None
        and school_limit_series is not None
    ):
        school_mask = ensure_series(school_limit_series).fillna(False).astype(bool)
        if school_mask.any():
            mentor_ids = mentor_id_series.astype("string").fillna("")
            sheet2.loc[school_mask, sheet_alias] = mentor_ids[school_mask]

    return sheet2


def build_summary_frame(
    exporter_cfg: Mapping[str, Any],
    *,
    total_students: int,
    allocated_count: int,
    error_count: int,
    dedupe_logs: Sequence[Mapping[str, Sequence[str]]] | None = None,
) -> pd.DataFrame:
    sheet_cfg = exporter_cfg.get("sheets", {}).get("Summary")
    if not isinstance(sheet_cfg, Mapping):
        return pd.DataFrame()
    columns = list(sheet_cfg.get("columns", []))
    if len(columns) < 2:
        return pd.DataFrame(columns=columns)
    data: list[dict[str, Any]] = []

    def _base_row(label: str, value: Any) -> dict[str, Any]:
        row = {columns[0]: label, columns[1]: value}
        for extra in columns[2:]:
            row[extra] = ""
        return row

    data.extend(
        [
            _base_row("تعداد کل دانش‌آموز", total_students),
            _base_row("تخصیص موفق", allocated_count),
            _base_row("تخصیص ناموفق", error_count),
        ]
    )
    if dedupe_logs:
        for entry in dedupe_logs:
            context = str(entry.get("context", "")).strip() or "dedupe"
            removed_columns = entry.get("columns") or []
            removed_text = ", ".join(str(col) for col in removed_columns)
            description = f"{context}: {removed_text}" if removed_text else context
            data.append(_base_row("پاکسازی ستون‌های تکراری", description))
    return pd.DataFrame(data, columns=columns)


def build_errors_frame(logs_df: pd.DataFrame | None, exporter_cfg: Mapping[str, Any]) -> pd.DataFrame:
    sheet_cfg = exporter_cfg.get("sheets", {}).get("Errors")
    if not isinstance(sheet_cfg, Mapping):
        return pd.DataFrame()
    columns = list(sheet_cfg.get("columns", []))
    if logs_df is None or logs_df.empty:
        return pd.DataFrame(columns=columns)
    status_series = logs_df.get("allocation_status")
    if status_series is None:
        return pd.DataFrame(columns=columns)
    fail_mask = status_series.astype("string") != "success"
    failed = logs_df.loc[fail_mask.fillna(False)]
    records = []
    for _, row in failed.iterrows():
        record = {
            columns[0]: row.get("student_id", ""),
        }
        if len(columns) > 1:
            record[columns[1]] = row.get("error_type") or row.get("allocation_status")
        if len(columns) > 2:
            record[columns[2]] = row.get("detailed_reason") or row.get("selection_reason")
        records.append(record)
    return pd.DataFrame(records, columns=columns)


def build_optional_sheet_frame(exporter_cfg: Mapping[str, Any], name: str) -> pd.DataFrame | None:
    sheet_cfg = exporter_cfg.get("sheets", {}).get(name)
    if not isinstance(sheet_cfg, Mapping):
        return None
    columns = list(sheet_cfg.get("columns", []))
    if not columns:
        return None
    return pd.DataFrame(columns=columns)


def ensure_template_workbook(template_path: str | Path, exporter_cfg: Mapping[str, Any]) -> Path:
    """ساخت یا به‌روزرسانی فایل قالب بر اساس تنظیمات موجود."""

    from openpyxl import Workbook, load_workbook

    path = Path(template_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    sheets_cfg = exporter_cfg.get("sheets", {}) if isinstance(exporter_cfg, Mapping) else {}

    def _expected_columns(sheet_cfg: Mapping[str, Any] | Sequence[Any]) -> list[str]:
        if isinstance(sheet_cfg, Mapping):
            columns = sheet_cfg.get("columns", [])
            if isinstance(columns, Mapping):
                return [str(key) for key in columns.keys()]
            if isinstance(columns, Sequence):
                return [str(value) for value in columns]
            return []
        if isinstance(sheet_cfg, Sequence):
            return [str(value) for value in sheet_cfg]
        return []

    def _write_headers(ws, columns: Sequence[str]) -> None:
        if not columns:
            return
        for col_idx, column_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=column_name)
        extra_cols = ws.max_column - len(columns)
        if extra_cols > 0:
            ws.delete_cols(len(columns) + 1, extra_cols)

    if path.exists():
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
        workbook.active.title = "Sheet2"

    for sheet_name, sheet_cfg in sheets_cfg.items():
        expected = _expected_columns(sheet_cfg)
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            ws = workbook.create_sheet(title=sheet_name)
        _write_headers(ws, expected)

    workbook.save(path)
    return path


def _write_dataframe_to_sheet(ws, df: pd.DataFrame) -> None:
    if df is None:
        return
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, value in enumerate(row.tolist(), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    ensure_text_columns(ws, df.columns)


def _normalize_header_label(value: Any) -> str:
    text = str(value or "").strip()
    text = text.translate(_DIGIT_TRANSLATION)
    text = _HEADER_PUNCT_PATTERN.sub(" ", text)
    text = normalize_fa(text)
    text = _HEADER_SPACE_PATTERN.sub(" ", text)
    return text.strip()


def _headers_equivalent(template: Sequence[str], expected: Sequence[str]) -> bool:
    normalized_template = [_normalize_header_label(value) for value in template]
    normalized_expected = [_normalize_header_label(value) for value in expected]
    return normalized_template == normalized_expected


def _rewrite_sheet_headers(ws, expected: Sequence[str]) -> None:
    for col_idx, value in enumerate(expected, start=1):
        ws.cell(row=1, column=col_idx, value=value)
    max_col = ws.max_column
    extra_cols = max_col - len(expected)
    if extra_cols > 0:
        ws.delete_cols(len(expected) + 1, extra_cols)


def _verify_headers(
    ws,
    expected: Sequence[str],
    *,
    on_mismatch: Callable[[str, Sequence[str], Sequence[str]], None] | None = None,
) -> None:
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [cell.value if cell.value is not None else "" for cell in header_cells]
    expected_list = list(expected)
    if len(headers) < len(expected_list):
        raise ValueError(
            f"Template sheet '{ws.title}' has fewer columns ({len(headers)}) than expected "
            f"({len(expected_list)})"
        )
    template_headers = headers[: len(expected_list)]
    if _headers_equivalent(template_headers, expected_list):
        return
    if on_mismatch is not None:
        on_mismatch(ws.title, template_headers, expected_list)
    print(f"⚠️  Rewriting headers in sheet '{ws.title}' to match config exactly")
    _rewrite_sheet_headers(ws, expected_list)


def build_header_signature(
    headers: Sequence[str], exporter_cfg: Mapping[str, Any] | None
) -> str:
    """محاسبهٔ امضای هدر بر پایه نسخه و لیست ستون‌ها.

    مثال
    ------
    >>> build_header_signature(["A", "B"], {"version": "1.1"})
    'v1.1-01ba4719c80b'
    """

    version = ""
    if isinstance(exporter_cfg, Mapping):
        version = str(exporter_cfg.get("version") or "").strip()
    version = version or "0"
    normalized = [_normalize_header_label(value) for value in headers]
    payload = "\n".join(normalized).encode("utf-8")
    digest = hashlib.sha256(payload).hexdigest()[:12]
    return f"v{version}-{digest}"


def _append_header_mismatch_row(summary_df: pd.DataFrame, log: HeaderMismatchLog) -> None:
    if not isinstance(summary_df, pd.DataFrame):
        return
    columns = list(summary_df.columns)
    if not columns:
        return
    row = {column: "" for column in columns}
    row[columns[0]] = "هشدار ناسازگاری هدر"
    if len(columns) > 1:
        row[columns[1]] = log.sheet_name
    detail_parts = [f"expected={log.expected_signature}"]
    if log.template_headers:
        template_text = " | ".join(log.template_headers)
        detail_parts.append(f"template={template_text}")
    target_column = columns[2] if len(columns) > 2 else columns[-1]
    row[target_column] = "؛ ".join(detail_parts)
    summary_df.loc[len(summary_df)] = row


def write_import_to_sabt_excel(
    df_sheet2: pd.DataFrame,
    df_summary: pd.DataFrame,
    df_errors: pd.DataFrame,
    df_sheet5: pd.DataFrame | None,
    df_9394: pd.DataFrame | None,
    template_path: str | Path,
    output_path: str | Path,
) -> None:
    """نوشتن خروجی ImportToSabt روی قالب موجود بدون تغییر استایل."""

    from openpyxl import load_workbook

    exporter_cfg = df_sheet2.attrs.get("exporter_config") or {}
    template = ensure_template_workbook(template_path, exporter_cfg)
    workbook = load_workbook(template)
    sheets_cfg = exporter_cfg.get("sheets", {}) if isinstance(exporter_cfg, Mapping) else {}
    if "Sheet2" not in workbook.sheetnames:
        raise ValueError("Template is missing sheet 'Sheet2'")
    header_logs = df_summary.attrs.get("header_mismatch_logs")
    if not isinstance(header_logs, list):
        header_logs = []
        df_summary.attrs["header_mismatch_logs"] = header_logs

    def _record_header_mismatch(
        sheet_name: str, template_headers: Sequence[str], expected_headers: Sequence[str]
    ) -> None:
        log = HeaderMismatchLog(
            sheet_name=sheet_name,
            template_headers=tuple(str(value or "") for value in template_headers),
            expected_headers=tuple(str(value or "") for value in expected_headers),
            expected_signature=build_header_signature(expected_headers, exporter_cfg),
        )
        header_logs.append(log)
        _append_header_mismatch_row(df_summary, log)

    sheets = [
        ("Sheet2", df_sheet2),
        ("Errors", df_errors),
        ("Sheet5", df_sheet5),
        ("9394", df_9394),
        ("Summary", df_summary),
    ]
    for name, df in sheets:
        if df is None:
            continue
        if name not in sheets_cfg and name not in workbook.sheetnames:
            continue
        if name not in workbook.sheetnames:
            ws = workbook.create_sheet(title=name)
        else:
            ws = workbook[name]
        _verify_headers(ws, df.columns, on_mismatch=_record_header_mismatch)
        _write_dataframe_to_sheet(ws, df)
    workbook.save(Path(output_path))
