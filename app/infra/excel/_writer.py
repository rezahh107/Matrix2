"""توابع کمکی برای کنترل قالب ستون‌های حساس در خروجی اکسل."""

from __future__ import annotations

from typing import Iterable, Sequence

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.common.contact_columns import TEXT_SENSITIVE_COLUMN_NAMES

__all__ = ["ensure_text_columns", "TEXT_COLUMN_NAMES"]

TEXT_COLUMN_NAMES = frozenset(TEXT_SENSITIVE_COLUMN_NAMES)


def ensure_text_columns(
    ws: Worksheet,
    headers: Sequence[str],
    *,
    extra_columns: Iterable[str] | None = None,
) -> None:
    """اعمال قالب «Text» برای ستون‌های شماره تلفن و کد رهگیری."""

    target_names = set(TEXT_COLUMN_NAMES)
    if extra_columns is not None:
        target_names.update(str(name) for name in extra_columns)
    header_list = [str(label) for label in headers]
    indexes = [idx for idx, label in enumerate(header_list) if label in target_names]
    if not indexes:
        return
    max_row = ws.max_row or 1
    for idx in indexes:
        column_letter = get_column_letter(idx + 1)
        for row in range(1, max_row + 1):
            cell = ws[f"{column_letter}{row}"]
            cell.number_format = "@"
