"""اعمال تنظیمات یکتای خروجی Excel (فونت، RTL، جدول‌ها)."""

from __future__ import annotations

import logging
from typing import Dict, Iterable, Tuple

import pandas as pd

from app.core.common.logging_ext import log_step

from .styles import build_font_config, ensure_openpyxl_named_style, ensure_xlsxwriter_format
from .tables import (
    TableNameRegistry,
    build_openpyxl_table,
    build_xlsxwriter_table,
    dedupe_headers,
)

__all__ = ["apply_workbook_formatting", "write_selection_reasons_sheet"]


_LOGGER = logging.getLogger(__name__)
_FONT_WARNING_EMITTED = False
_SCHEMA_DEFINED_NAME = "__SELECTION_REASON_SCHEMA_HASH__"


def _warn_fonts_not_embedded() -> None:
    global _FONT_WARNING_EMITTED
    if not _FONT_WARNING_EMITTED:
        _LOGGER.warning(
            "فونت‌های سفارشی در فایل‌های Excel جاسازی نمی‌شوند؛ برای اشتراک‌گذاری امن، خروجی را به PDF تبدیل کنید."
        )
        _FONT_WARNING_EMITTED = True


def _format_xlsxwriter(
    writer: pd.ExcelWriter,
    sheet_frames: Dict[str, pd.DataFrame],
    *,
    rtl: bool,
    font_name: str | None,
    font_size: int | None,
) -> None:
    workbook = writer.book  # type: ignore[attr-defined]
    worksheet_map = writer.sheets  # type: ignore[attr-defined]
    font = build_font_config(font_name, font_size=font_size)
    body_fmt = ensure_xlsxwriter_format(workbook, font)
    right_body_fmt = ensure_xlsxwriter_format(workbook, font, align_right=True)
    header_fmt = ensure_xlsxwriter_format(workbook, font, header=True, align_right=rtl)
    table_names = TableNameRegistry()

    for sheet_name, df in sheet_frames.items():
        worksheet = worksheet_map[sheet_name]
        if rtl:
            worksheet.right_to_left()
        worksheet.freeze_panes(1, 0)
        column_count = len(df.columns)
        if column_count:
            worksheet.set_row(0, None, header_fmt)
            datetime_columns = {
                column
                for column, dtype in df.dtypes.items()
                if pd.api.types.is_datetime64_any_dtype(dtype)
            }
            string_columns = {
                idx for idx, dtype in enumerate(df.dtypes, start=0) if not pd.api.types.is_numeric_dtype(dtype)
            }
            for idx, column in enumerate(df.columns):
                if column in datetime_columns:
                    continue
                fmt = right_body_fmt if idx in string_columns else body_fmt
                worksheet.set_column(idx, idx, None, fmt)
        if df.empty or column_count == 0:
            continue
        table_name = table_names.reserve(sheet_name)
        last_row = len(df)
        last_col = df.shape[1] - 1
        worksheet.add_table(0, 0, last_row, last_col, build_xlsxwriter_table(df, table_name))


def _format_openpyxl(
    writer: pd.ExcelWriter,
    sheet_frames: Dict[str, pd.DataFrame],
    *,
    rtl: bool,
    font_name: str | None,
    font_size: int | None,
) -> None:
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    workbook = writer.book  # type: ignore[attr-defined]
    font = build_font_config(font_name, font_size=font_size)
    style_name = ensure_openpyxl_named_style(workbook, font)
    table_names = TableNameRegistry()

    for sheet_name, df in sheet_frames.items():
        worksheet = workbook[sheet_name]
        if rtl:
            worksheet.sheet_view.rightToLeft = True
        worksheet.freeze_panes = "A2"
        max_col = max(len(df.columns), worksheet.max_column)
        max_row = max(len(df) + 1, worksheet.max_row)
        right_aligned = {
            idx
            for idx, dtype in enumerate(df.dtypes, start=1)
            if not pd.api.types.is_numeric_dtype(dtype)
        }
        if max_col and max_row:
            for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.style = style_name
                    if cell.col_idx in right_aligned:
                        cell.alignment = Alignment(horizontal="right")
        if df.empty or df.shape[1] == 0:
            continue
        headers = dedupe_headers(df.columns)
        table_name = table_names.reserve(sheet_name)
        ref = f"A1:{get_column_letter(df.shape[1])}{len(df) + 1}"
        worksheet.add_table(build_openpyxl_table(table_name, ref, headers))
        for idx, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=idx, value=header)


def apply_workbook_formatting(
    writer: pd.ExcelWriter,
    *,
    engine: str,
    sheet_frames: Dict[str, pd.DataFrame],
    rtl: bool,
    font_name: str | None,
    font_size: int | None,
) -> None:
    """اعمال تنظیمات خروجی Excel پس از نوشتن DataFrameها.

    مثال::

        >>> with pd.ExcelWriter("/tmp/out.xlsx", engine="openpyxl") as writer:
        ...     df = pd.DataFrame({"A": [1, 2]})
        ...     df.to_excel(writer, sheet_name="Sheet", index=False)
        ...     apply_workbook_formatting(
        ...         writer,
        ...         engine="openpyxl",
        ...         sheet_frames={"Sheet": df},
        ...         rtl=True,
        ...         font_name="Vazirmatn",
        ...     )
    """

    _warn_fonts_not_embedded()

    if engine == "xlsxwriter":
        _format_xlsxwriter(
            writer,
            sheet_frames,
            rtl=rtl,
            font_name=font_name,
            font_size=font_size,
        )
    elif engine == "openpyxl":
        _format_openpyxl(
            writer,
            sheet_frames,
            rtl=rtl,
            font_name=font_name,
            font_size=font_size,
        )


def _sanitize_selection_reasons_frame(
    df_reasons: pd.DataFrame | None,
    columns: Iterable[str],
) -> pd.DataFrame:
    """پاک‌سازی DataFrame دلایل انتخاب متناسب با ستون‌های Policy.

    مثال::

        >>> import pandas as pd
        >>> raw = pd.DataFrame({"نام": ["الف"]})
        >>> _sanitize_selection_reasons_frame(raw, ("شمارنده", "نام"))
           شمارنده   نام
        0         1  الف
    """

    ordered_columns: Tuple[str, ...] = tuple(columns)
    if df_reasons is None or df_reasons.empty:
        sanitized = pd.DataFrame(columns=ordered_columns)
    else:
        sanitized = df_reasons.copy()

    for column in ordered_columns:
        if column not in sanitized.columns:
            sanitized[column] = pd.NA

    sanitized = sanitized.loc[:, ordered_columns].copy()
    counter = pd.Series(
        range(1, len(sanitized) + 1),
        index=sanitized.index,
        dtype="Int64",
    )
    sanitized.loc[:, "شمارنده"] = counter

    str_columns = [column for column in ordered_columns if column != "شمارنده"]
    if str_columns:
        sanitized = sanitized.astype({column: "string" for column in str_columns})

    return sanitized


def _set_schema_hash_defined_name(writer: pd.ExcelWriter | None, schema_hash: str) -> None:
    """ثبت Schema Hash به‌صورت Defined Name در فایل Excel.

    مثال::

        >>> import pandas as pd
        >>> with pd.ExcelWriter("/tmp/sample.xlsx", engine="xlsxwriter") as writer:
        ...     _set_schema_hash_defined_name(writer, "abc123")
        ...     writer.book.defined_names_dict["__SELECTION_REASON_SCHEMA_HASH__"]
        ('__SELECTION_REASON_SCHEMA_HASH__', '="abc123"')
    """

    if writer is None or not schema_hash:
        return

    workbook = getattr(writer, "book", None)
    engine = str(getattr(writer, "engine", "") or "").lower()
    if workbook is None:
        return

    if engine == "openpyxl":
        from openpyxl.workbook.defined_name import DefinedName

        defined_names = workbook.defined_names
        if _SCHEMA_DEFINED_NAME in defined_names:
            del defined_names[_SCHEMA_DEFINED_NAME]
        defined_names.add(
            DefinedName(name=_SCHEMA_DEFINED_NAME, attr_text=f'"{schema_hash}"')
        )
    elif engine == "xlsxwriter":
        workbook.define_name(_SCHEMA_DEFINED_NAME, f'="{schema_hash}"')


def write_selection_reasons_sheet(
    df_reasons: pd.DataFrame | None,
    writer: pd.ExcelWriter | None,
    policy: "PolicyConfig",
) -> tuple[str, pd.DataFrame]:
    """تهیه و نوشتن شیت «دلایل انتخاب پشتیبان» مطابق Policy.

    مثال:
        >>> from app.core.policy_loader import load_policy
        >>> policy = load_policy()
        >>> df = pd.DataFrame({"کدملی": ["001"], "دلیل انتخاب پشتیبان": ["آزمایشی"]})
        >>> sheet, sanitized = write_selection_reasons_sheet(df, writer=None, policy=policy)
        >>> sheet
        'دلایل انتخاب پشتیبان'
        >>> sanitized.loc[0, "شمارنده"]
        1
    """

    from app.core.policy_loader import PolicyConfig  # محلی برای جلوگیری از حلقهٔ import

    if not isinstance(policy, PolicyConfig):  # pragma: no cover - نگهبان برای تزریق نادرست
        raise TypeError("policy must be an instance of PolicyConfig")

    options = policy.emission.selection_reasons
    sheet_name = options.sheet_name
    columns = options.columns
    with log_step(_LOGGER, "selection_reason_export"):
        sanitized = _sanitize_selection_reasons_frame(df_reasons, columns)
        sanitized.attrs["schema_hash"] = options.schema_hash

        if writer is not None:
            sanitized.to_excel(writer, sheet_name=sheet_name, index=False)
            _set_schema_hash_defined_name(writer, options.schema_hash)

    return sheet_name, sanitized
