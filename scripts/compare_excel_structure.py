#!/usr/bin/env python3
"""اسکریپت مقایسه ساختار دو فایل اکسل."""

from __future__ import annotations

import sys
from difflib import unified_diff
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


def normalize_header(value: object) -> str:
    """نرمال‌سازی متن هدر برای مقایسه خطاها."""

    if value is None:
        return ""
    text = str(value).strip()
    import re

    text = re.sub(r"\s+", " ", text)
    return text


def _trim_trailing_empty(headers: list[str]) -> list[str]:
    while headers and headers[-1] == "":
        headers.pop()
    return headers


def compare_excel_structure(file1_path: Path, file2_path: Path) -> bool:
    """مقایسه ساختار دو فایل اکسل و چاپ تفاوت‌ها."""

    print("📊 Comparing Excel structures:")
    print(f"  File 1: {file1_path}")
    print(f"  File 2: {file2_path}\n")

    wb1 = load_workbook(file1_path, read_only=True)
    wb2 = load_workbook(file2_path, read_only=True)

    sheets1 = set(wb1.sheetnames)
    sheets2 = set(wb2.sheetnames)

    if sheets1 != sheets2:
        print("❌ Sheet names differ!")
        print(f"  Only in file1: {sheets1 - sheets2}")
        print(f"  Only in file2: {sheets2 - sheets1}")
        return False

    print(f"✅ Both files have {len(sheets1)} sheets: {', '.join(sorted(sheets1))}\n")
    all_match = True

    for sheet_name in sorted(sheets1):
        print(f"📄 Sheet: {sheet_name}")
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]
        headers1 = _trim_trailing_empty(
            [normalize_header(cell.value) for cell in next(ws1.iter_rows(min_row=1, max_row=1))]
        )
        headers2 = _trim_trailing_empty(
            [normalize_header(cell.value) for cell in next(ws2.iter_rows(min_row=1, max_row=1))]
        )
        if headers1 == headers2:
            print(f"  ✅ Headers match ({len(headers1)} columns)\n")
            continue
        print("  ❌ Headers differ!")
        print(f"    File1 has {len(headers1)} columns")
        print(f"    File2 has {len(headers2)} columns")
        diff = list(
            unified_diff(
                headers1,
                headers2,
                fromfile=f"{file1_path.name}/{sheet_name}",
                tofile=f"{file2_path.name}/{sheet_name}",
                lineterm="",
            )
        )
        if diff:
            print("\n  Differences:")
            for line in diff[:20]:
                print(f"    {line}")
            if len(diff) > 20:
                print(f"    ... ({len(diff) - 20} more lines)")
        print()
        all_match = False

    wb1.close()
    wb2.close()

    if all_match:
        print("🎉 All structures match!")
    else:
        print("❌ Structures do not match")
    return all_match


def _usage() -> str:
    return "Usage: python compare_excel_structure.py <file1.xlsx> <file2.xlsx>"


def _validate_paths(paths: Iterable[Path]) -> None:
    for path in paths:
        if not path.exists():
            raise FileNotFoundError(f"Error: {path} does not exist")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(_usage())
        sys.exit(1)

    file1 = Path(sys.argv[1])
    file2 = Path(sys.argv[2])

    try:
        _validate_paths((file1, file2))
    except FileNotFoundError as exc:
        print(exc)
        sys.exit(1)

    match = compare_excel_structure(file1, file2)
    sys.exit(0 if match else 1)
